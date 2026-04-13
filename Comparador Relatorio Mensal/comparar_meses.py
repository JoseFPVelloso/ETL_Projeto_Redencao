import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import numpy as np
import os
import re
from datetime import datetime

# ── Diretório do script ───────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Cores ─────────────────────────────────────────────────────────────────────
COR_BG       = "#1e1e2e"
COR_PAINEL   = "#2a2a3e"
COR_BORDA    = "#3a3a5e"
COR_ACENTO   = "#7c6af7"
COR_TEXTO    = "#e0e0f0"
COR_SUBTEXTO = "#9090b0"
COR_ALTA     = "#4ade80"
COR_QUEDA    = "#f87171"
COR_BTN      = "#5a4fcf"
COR_BTN_HOV  = "#7c6af7"

# ── Mapeamento de meses em português ─────────────────────────────────────────
MESES_PT = {
    "janeiro": 1, "fevereiro": 2, "marco": 3, "março": 3, "abril": 4,
    "maio": 5, "junho": 6, "julho": 7, "agosto": 8,
    "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12,
}
MESES_NUM_PT = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}


def normalizar_data(valor):
    """
    Converte qualquer valor de data/mes para datetime(ano, mes, 1).
    Aceita: Timestamp com hora, "Janeiro, 2026", "2026-01-01 00:00:00", etc.
    Retorna datetime ou None.
    """
    if pd.isna(valor) if not isinstance(valor, str) else False:
        return None

    # Ja e datetime / Timestamp
    if isinstance(valor, (pd.Timestamp, datetime)):
        ts = pd.Timestamp(valor)
        return datetime(ts.year, ts.month, 1)

    s = str(valor).strip()

    # Formato "Mes, Ano" em portugues (ex: "Marco, 2026" ou "Março, 2026")
    match = re.match(r"([A-Za-zÀ-ú]+)[,\s]+(\d{4})", s)
    if match:
        nome_mes = match.group(1).lower().strip()
        # remover acento simples para comparacao
        nome_mes_norm = nome_mes.replace("ç", "c").replace("ã", "a")
        ano = int(match.group(2))
        num = MESES_PT.get(nome_mes) or MESES_PT.get(nome_mes_norm)
        if num:
            return datetime(ano, num, 1)

    # Tenta parsear com pandas
    try:
        ts = pd.to_datetime(s, dayfirst=True)
        return datetime(ts.year, ts.month, 1)
    except Exception:
        pass

    return None


def formatar_label(dt):
    """Converte datetime -> 'Mes, Ano' para exibir no combo."""
    return f"{MESES_NUM_PT[dt.month]}, {dt.year}"


def detectar_coluna_data(df):
    candidatas = ["Data", "data", "Mes", "Mês", "mes", "mês", "Período", "periodo"]
    for c in candidatas:
        if c in df.columns:
            return c
    for c in df.columns:
        sample = df[c].dropna().head(10)
        acertos = sum(1 for v in sample if normalizar_data(v) is not None)
        if acertos >= max(1, len(sample) * 0.6):
            return c
    return None


# ── Estado global ─────────────────────────────────────────────────────────────
estado = {
    "df": None,
    "filepath": None,
    "col_data": None,
    "meses_dt": [],
    "meses_label": [],
    "resultado": [],
    "mes1": "",
    "mes2": "",
}


def carregar_arquivo():
    path = filedialog.askopenfilename(
        title="Selecionar planilha",
        filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")]
    )
    if not path:
        return
    try:
        df = pd.read_excel(path)
        estado["df"] = df
        estado["filepath"] = path
        nome = os.path.basename(path)
        lbl_arquivo.config(text=f"📄  {nome}", fg=COR_ALTA)

        col_data = detectar_coluna_data(df)
        if not col_data:
            messagebox.showerror("Erro", "Não foi possível identificar uma coluna de data/mês.")
            return
        estado["col_data"] = col_data

        # Normalizar todas as datas únicas e agrupar por (ano, mes)
        datas_raw = df[col_data].dropna().unique()
        seen = {}
        for v in datas_raw:
            dt = normalizar_data(v)
            if dt and dt not in seen:
                seen[dt] = v

        ordenados = sorted(seen.keys())
        labels = [formatar_label(dt) for dt in ordenados]

        estado["meses_dt"]    = ordenados
        estado["meses_label"] = labels

        combo_mes1["values"] = labels
        combo_mes2["values"] = labels
        if len(labels) >= 2:
            combo_mes1.set(labels[-2])
            combo_mes2.set(labels[-1])
        elif len(labels) == 1:
            combo_mes1.set(labels[0])

        frame_combos.pack(pady=(0, 12))
        btn_comparar.pack(pady=(0, 10))
        lbl_status.config(
            text=f"✔  {len(labels)} período(s) detectado(s) | coluna: '{col_data}'",
            fg=COR_SUBTEXTO
        )

    except Exception as e:
        messagebox.showerror("Erro ao carregar", str(e))


def comparar():
    label1 = combo_mes1.get().strip()
    label2 = combo_mes2.get().strip()
    if not label1 or not label2:
        messagebox.showwarning("Atenção", "Selecione os dois períodos.")
        return
    if label1 == label2:
        messagebox.showwarning("Atenção", "Selecione períodos diferentes.")
        return

    labels = estado["meses_label"]
    dts    = estado["meses_dt"]
    dt1    = dts[labels.index(label1)]
    dt2    = dts[labels.index(label2)]

    df  = estado["df"]
    col = estado["col_data"]

    mask1 = df[col].apply(lambda v: normalizar_data(v) == dt1)
    mask2 = df[col].apply(lambda v: normalizar_data(v) == dt2)
    d1 = df[mask1]
    d2 = df[mask2]

    if d1.empty or d2.empty:
        messagebox.showerror("Erro", "Um dos períodos não tem dados.")
        return

    num_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    s1 = d1[num_cols].sum()
    s2 = d2[num_cols].sum()

    try:
        threshold = float(entry_threshold.get() or 10)
    except ValueError:
        threshold = 10.0

    rows = []
    for c in num_cols:
        v1, v2 = s1[c], s2[c]
        if pd.isna(v1) or pd.isna(v2) or v1 == 0:
            continue
        var = (v2 - v1) / abs(v1) * 100
        if abs(var) >= threshold:
            rows.append((c, v1, v2, var))

    rows.sort(key=lambda x: x[3], reverse=True)

    for item in tree.get_children():
        tree.delete(item)

    for indicador, v1, v2, var in rows:
        sinal = "▲" if var > 0 else "▼"
        tag   = "alta" if var > 0 else "queda"
        tree.insert("", "end",
                    values=(indicador, f"{v1:,.0f}", f"{v2:,.0f}", f"{sinal} {var:+.1f}%"),
                    tags=(tag,))

    n      = len(rows)
    altas  = sum(1 for _, _, _, v in rows if v > 0)
    quedas = n - altas
    lbl_status.config(
        text=f"  {n} indicador(es) com variação ≥ {threshold:.0f}%  •  {altas} altas  •  {quedas} quedas",
        fg=COR_SUBTEXTO
    )
    lbl_cabecalho.config(text=f"Comparativo:  {label1}  →  {label2}", fg=COR_TEXTO)
    frame_resultado.pack(fill="both", expand=True, padx=20, pady=(0, 16))

    estado["resultado"] = rows
    estado["mes1"]      = label1
    estado["mes2"]      = label2
    btn_exportar.config(state="normal")


def exportar():
    rows = estado["resultado"]
    if not rows:
        return

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    mes1 = estado["mes1"]
    mes2 = estado["mes2"]

    nome_safe = f"comparativo_{mes1}_{mes2}".replace(", ", "_").replace(" ", "_")
    path_out  = os.path.join(SCRIPT_DIR, f"{nome_safe}.xlsx")

    try:
        threshold = float(entry_threshold.get() or 10)
    except ValueError:
        threshold = 10.0

    altas  = sum(1 for _, _, _, v in rows if v > 0)
    quedas = sum(1 for _, _, _, v in rows if v < 0)

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparativo"

    borda = Border(
        bottom=Side(style="thin", color="4A4A7A"),
    )

    # ── Linha 1: Título ───────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value     = f"Comparativo de Indicadores: {mes1} → {mes2}"
    c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=13)
    c.fill      = PatternFill("solid", start_color="3B3B6B")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Linha 2: Resumo ───────────────────────────────────────────────────────
    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value     = f"Variação mínima: {threshold:.0f}%   |   {altas} altas   •   {quedas} quedas"
    c.font      = Font(italic=True, color="9090C0", name="Arial", size=9)
    c.fill      = PatternFill("solid", start_color="22223A")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Linha 3: Cabeçalho ────────────────────────────────────────────────────
    headers = ["Indicador", mes1, mes2, "Variação (%)"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col_idx)
        c.value     = h
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill      = PatternFill("solid", start_color="4A4A8A")
        c.alignment = Alignment(horizontal="left" if col_idx == 1 else "center", vertical="center")
        c.border    = borda
    ws.row_dimensions[3].height = 22

    # ── Linhas de dados ───────────────────────────────────────────────────────
    for i, (indicador, v1, v2, var) in enumerate(rows, 4):
        alta      = var > 0
        sinal     = "▲" if alta else "▼"
        fill_par  = PatternFill("solid", start_color="1C2C1C" if alta else "2C1C1C")
        fill_impar= PatternFill("solid", start_color="1A2A1A" if alta else "2A1A1A")
        fill_row  = fill_par if i % 2 == 0 else fill_impar
        fill_var  = PatternFill("solid", start_color="14381A" if alta else "381414")
        cor_var   = "4ADE80" if alta else "F87171"

        c1 = ws.cell(row=i, column=1, value=indicador)
        c1.font      = Font(name="Arial", size=10, color="E0E0F0")
        c1.fill      = fill_row
        c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        c2 = ws.cell(row=i, column=2, value=round(v1, 1))
        c2.font          = Font(name="Arial", size=10, color="B0B0D0")
        c2.fill          = fill_row
        c2.alignment     = Alignment(horizontal="right", vertical="center")
        c2.number_format = "#,##0.0"

        c3 = ws.cell(row=i, column=3, value=round(v2, 1))
        c3.font          = Font(name="Arial", size=10, color="B0B0D0")
        c3.fill          = fill_row
        c3.alignment     = Alignment(horizontal="right", vertical="center")
        c3.number_format = "#,##0.0"

        c4 = ws.cell(row=i, column=4, value=f"{sinal} {var:+.1f}%")
        c4.font      = Font(bold=True, name="Arial", size=10, color=cor_var)
        c4.fill      = fill_var
        c4.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[i].height = 20

    # ── Larguras e freeze ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.freeze_panes = "A4"

    wb.save(path_out)
    messagebox.showinfo("✔  Exportado com sucesso", f"Arquivo salvo em:\n{path_out}")


# ── Janela principal ──────────────────────────────────────────────────────────
root = tk.Tk()
root.title("Comparador de Períodos — Programa Redenção")
root.configure(bg=COR_BG)
root.geometry("880x700")
root.minsize(720, 520)

style = ttk.Style()
style.theme_use("clam")
style.configure("TCombobox",
    fieldbackground=COR_PAINEL, background=COR_PAINEL,
    foreground=COR_TEXTO, selectbackground=COR_ACENTO,
    bordercolor=COR_BORDA, arrowcolor=COR_TEXTO)
style.configure("Treeview",
    background=COR_PAINEL, foreground=COR_TEXTO,
    fieldbackground=COR_PAINEL, rowheight=28,
    bordercolor=COR_BORDA, font=("Arial", 10))
style.configure("Treeview.Heading",
    background=COR_BORDA, foreground=COR_TEXTO,
    font=("Arial", 10, "bold"), relief="flat")
style.map("Treeview", background=[("selected", COR_ACENTO)])

# ── Header ────────────────────────────────────────────────────────────────────
frame_header = tk.Frame(root, bg=COR_PAINEL, pady=18)
frame_header.pack(fill="x")
tk.Label(frame_header, text="📊  Comparador de Períodos",
         bg=COR_PAINEL, fg=COR_TEXTO, font=("Arial", 16, "bold")).pack()
tk.Label(frame_header, text="Carregue uma planilha e compare dois meses",
         bg=COR_PAINEL, fg=COR_SUBTEXTO, font=("Arial", 10)).pack(pady=(2, 0))

# ── Controles ─────────────────────────────────────────────────────────────────
frame_ctrl = tk.Frame(root, bg=COR_BG, pady=16)
frame_ctrl.pack(fill="x", padx=20)

def on_enter(e): e.widget.config(bg=COR_BTN_HOV)
def on_leave(e): e.widget.config(bg=COR_BTN)

btn_carregar = tk.Button(frame_ctrl, text="  📂  Carregar Planilha  ",
    bg=COR_BTN, fg=COR_TEXTO, font=("Arial", 11, "bold"),
    relief="flat", cursor="hand2", padx=10, pady=8, command=carregar_arquivo)
btn_carregar.pack()
btn_carregar.bind("<Enter>", on_enter)
btn_carregar.bind("<Leave>", on_leave)

lbl_arquivo = tk.Label(frame_ctrl, text="Nenhum arquivo carregado",
    bg=COR_BG, fg=COR_SUBTEXTO, font=("Arial", 9))
lbl_arquivo.pack(pady=(6, 0))

# Combos de meses
frame_combos = tk.Frame(frame_ctrl, bg=COR_BG)

tk.Label(frame_combos, text="Mês anterior (Passado):", bg=COR_BG, fg=COR_SUBTEXTO,
         font=("Arial", 9)).grid(row=0, column=0, padx=(0, 6), sticky="w")
combo_mes1 = ttk.Combobox(frame_combos, width=22, state="readonly", font=("Arial", 10))
combo_mes1.grid(row=0, column=1, padx=(0, 20))

tk.Label(frame_combos, text="Mês atual (Presente/Futuro):", bg=COR_BG, fg=COR_SUBTEXTO,
         font=("Arial", 9)).grid(row=0, column=2, padx=(0, 6), sticky="w")
combo_mes2 = ttk.Combobox(frame_combos, width=22, state="readonly", font=("Arial", 10))
combo_mes2.grid(row=0, column=3, padx=(0, 20))

tk.Label(frame_combos, text="Variação mínima (%):", bg=COR_BG, fg=COR_SUBTEXTO,
         font=("Arial", 9)).grid(row=0, column=4, padx=(0, 6), sticky="w")
entry_threshold = tk.Entry(frame_combos, width=5, bg=COR_PAINEL, fg=COR_TEXTO,
    insertbackground=COR_TEXTO, font=("Arial", 10), relief="flat",
    highlightthickness=1, highlightbackground=COR_BORDA)
entry_threshold.insert(0, "10")
entry_threshold.grid(row=0, column=5)

frame_combos.pack_forget()

btn_comparar = tk.Button(frame_ctrl, text="  🔍  Comparar  ",
    bg=COR_ACENTO, fg="white", font=("Arial", 11, "bold"),
    relief="flat", cursor="hand2", padx=10, pady=7, command=comparar)
btn_comparar.pack_forget()

lbl_status = tk.Label(frame_ctrl, text="", bg=COR_BG, fg=COR_SUBTEXTO, font=("Arial", 9))
lbl_status.pack(pady=(6, 0))

# ── Resultado ─────────────────────────────────────────────────────────────────
frame_resultado = tk.Frame(root, bg=COR_BG)

lbl_cabecalho = tk.Label(frame_resultado, text="", bg=COR_BG,
    fg=COR_TEXTO, font=("Arial", 11, "bold"))
lbl_cabecalho.pack(anchor="w", pady=(0, 8))

frame_tree = tk.Frame(frame_resultado, bg=COR_BORDA, bd=1)
frame_tree.pack(fill="both", expand=True)

scrollbar = ttk.Scrollbar(frame_tree, orient="vertical")
scrollbar.pack(side="right", fill="y")

cols = ("Indicador", "Mês base", "Mês comparado", "Variação")
tree = ttk.Treeview(frame_tree, columns=cols, show="headings",
                    yscrollcommand=scrollbar.set)
scrollbar.config(command=tree.yview)

tree.heading("Indicador",     text="Indicador",     anchor="w")
tree.heading("Mês base",      text="Mês base",      anchor="e")
tree.heading("Mês comparado", text="Mês comparado", anchor="e")
tree.heading("Variação",      text="Variação (%)",  anchor="center")

tree.column("Indicador",     width=390, anchor="w")
tree.column("Mês base",      width=110, anchor="e")
tree.column("Mês comparado", width=120, anchor="e")
tree.column("Variação",      width=110, anchor="center")

tree.tag_configure("alta",  background="#1a2e1a", foreground=COR_ALTA)
tree.tag_configure("queda", background="#2e1a1a", foreground=COR_QUEDA)
tree.pack(fill="both", expand=True)

frame_btn_exp = tk.Frame(frame_resultado, bg=COR_BG)
frame_btn_exp.pack(fill="x", pady=(10, 0))

btn_exportar = tk.Button(frame_btn_exp,
    text="  💾  Gerar Excel na pasta do script  ",
    bg="#1e5c3a", fg="white", font=("Arial", 10, "bold"),
    relief="flat", cursor="hand2", padx=12, pady=7,
    command=exportar, state="disabled")
btn_exportar.pack(side="right")

frame_resultado.pack_forget()

root.mainloop()