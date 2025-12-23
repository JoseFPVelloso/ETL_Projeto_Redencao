# Gerador de Relatório Consolidado por Período e Dia
# Replica funcionalidade do Google Sheets para Python
# Cria planilha com contagem diária formatada e análises
# Versão 2: Com cálculo automático de média anterior

# %% [markdown]
# # 1. Configuração Inicial

# %%
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path
import warnings
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# Definir caminho da raiz do projeto
if '__file__' in globals():
    projeto_root = Path(__file__).resolve().parent.parent
else:
    projeto_root = Path.cwd()
    if projeto_root.name == 'notebooks':
        projeto_root = projeto_root.parent
    if not (projeto_root / 'data').exists():
        projeto_root = projeto_root.parent

print("=" * 80)
print("GERADOR DE RELATÓRIO CONSOLIDADO POR PERÍODO E DIA")
print("=" * 80)
print(f"✓ Bibliotecas importadas")
print(f"✓ Processamento iniciado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

# %% [markdown]
# # 2. Constantes e Configurações

# %%
LIMIAR = 10  # Oculta ruas com valores <= 10 em todo o período
PROCESSED_DIR = projeto_root / 'data' / 'processed'
DOCS_DIR = projeto_root / 'docs'

DOCS_DIR.mkdir(exist_ok=True)

print(f"\n✓ Configurações:")
print(f"  • Limiar de aglomeração: > {LIMIAR} pessoas")
print(f"  • Raiz do projeto: {projeto_root}")
print(f"  • Diretório processados: {PROCESSED_DIR}")
print(f"  • Diretório saída: {DOCS_DIR}")

# %% [markdown]
# # 3. Funções Utilitárias

# %%
def normalizar_periodo(periodo_str):
    """Normaliza string de período para categoria padrão"""
    if not periodo_str:
        return None
    
    s = str(periodo_str).lower()
    if 'madrug' in s:
        return 'madrugada'
    if 'manh' in s or 'manha' in s:
        return 'manhã'
    if 'tarde' in s:
        return 'tarde'
    if 'noite' in s:
        return 'noite'
    return None

def gerar_lista_dias(data_inicio, data_fim):
    """Gera lista de dias entre duas datas"""
    dias = []
    data_atual = data_inicio
    while data_atual <= data_fim:
        dias.append(data_atual)
        data_atual += timedelta(days=1)
    return dias

def extrair_numero_logradouro(numero_str):
    """Extrai número inteiro do campo numero_logradouro"""
    if pd.isna(numero_str) or numero_str == '':
        return 999999  # Coloca no final os sem número
    
    # Tentar converter para int
    try:
        # Remover letras e caracteres especiais, pegar só os dígitos
        import re
        digitos = re.search(r'\d+', str(numero_str))
        if digitos:
            return int(digitos.group())
        return 999999
    except:
        return 999999

def ordenar_logradouros_df(df_input):
    """Ordena DataFrame por tipo_logradouro, nome_logradouro (alfabético) e numero_logradouro (numérico)"""
    df_ordenado = df_input.copy()
    
    # Criar coluna auxiliar com número extraído
    df_ordenado['_numero_ordem'] = df_ordenado['numero_logradouro'].apply(extrair_numero_logradouro)
    
    # Ordenar por:
    # 1. tipo_logradouro (alfabético) - ex: Alameda, Avenida, Rua
    # 2. nome_logradouro (alfabético) - ex: Barão de Piracicaba
    # 3. _numero_ordem (numérico) - ex: 57, 75, 431
    # 4. logradouro (desempate) - ex: caso haja duplicatas
    df_ordenado = df_ordenado.sort_values(
        by=['tipo_logradouro', 'nome_logradouro', '_numero_ordem', 'logradouro'],
        key=lambda x: x.str.lower() if x.dtype == "object" else x
    )
    
    # Remover coluna auxiliar
    df_ordenado = df_ordenado.drop(columns=['_numero_ordem'])
    
    return df_ordenado

print("✓ Funções utilitárias definidas")

# %% [markdown]
# # 4. Selecionar Planilha Processada

# %%
arquivos_disponiveis = sorted(PROCESSED_DIR.glob('*.xlsx'))

if not arquivos_disponiveis:
    print("\n⚠️  ERRO: Nenhum arquivo .xlsx encontrado na pasta 'data/processed/'")
    exit()

print(f"\n📁 Arquivos disponíveis em 'data/processed/':")
print("-" * 80)
for idx, arquivo in enumerate(arquivos_disponiveis, 1):
    print(f"  [{idx}] {arquivo.name}")
print("-" * 80)

while True:
    try:
        selecao = input(f"\nSelecione o número do arquivo [1-{len(arquivos_disponiveis)}]: ").strip()
        idx_selecionado = int(selecao) - 1
        
        if 0 <= idx_selecionado < len(arquivos_disponiveis):
            arquivo_selecionado = arquivos_disponiveis[idx_selecionado]
            print(f"\n✓ Arquivo selecionado: {arquivo_selecionado.name}")
            break
        else:
            print(f"⚠️  Por favor, escolha um número entre 1 e {len(arquivos_disponiveis)}")
    except ValueError:
        print("⚠️  Entrada inválida. Digite apenas o número do arquivo.")
    except KeyboardInterrupt:
        print("\n\n⚠️  Operação cancelada pelo usuário.")
        exit()

# %% [markdown]
# # 5. Solicitar Período do Relatório

# %%
print("\n" + "=" * 80)
print("DEFINIR PERÍODO DO RELATÓRIO")
print("=" * 80)

# Data início
while True:
    try:
        data_inicio_input = input("\n📅 Digite a DATA INICIAL (DD/MM/AAAA): ").strip()
        data_inicio = datetime.strptime(data_inicio_input, '%d/%m/%Y')
        break
    except ValueError:
        print("⚠️  Formato inválido. Use DD/MM/AAAA")

# Data fim
while True:
    try:
        data_fim_input = input("📅 Digite a DATA FINAL (DD/MM/AAAA): ").strip()
        data_fim = datetime.strptime(data_fim_input, '%d/%m/%Y')
        
        if data_fim < data_inicio:
            print("⚠️  A data final deve ser maior ou igual à data inicial")
            continue
        break
    except ValueError:
        print("⚠️  Formato inválido. Use DD/MM/AAAA")

print(f"\n✓ Período definido: {data_inicio.strftime('%d/%m/%Y')} até {data_fim.strftime('%d/%m/%Y')}")

# %% [markdown]
# # 6. Carregar e Preparar Dados

# %%
print(f"\n📊 Carregando dados...")
print("-" * 80)

try:
    df = pd.read_excel(arquivo_selecionado)
    print(f"✓ Planilha carregada: {len(df):,} registros")
except Exception as e:
    print(f"\n⚠️  ERRO ao carregar planilha: {e}")
    exit()

# Validar colunas necessárias (nomes da planilha processada)
colunas_necessarias = ['Data', 'Período', 'Qtd. pessoas', 'Logradouro', 'tipo_logradouro', 'nome_logradouro', 'numero_logradouro']
colunas_faltantes = [col for col in colunas_necessarias if col not in df.columns]

if colunas_faltantes:
    print(f"\n⚠️  ERRO: Colunas obrigatórias não encontradas: {colunas_faltantes}")
    print(f"\n📋 Colunas disponíveis na planilha:")
    for col in df.columns:
        print(f"  • {col}")
    exit()

# Padronizar nomes das colunas
df = df.rename(columns={
    'Data': 'data',
    'Período': 'periodo',
    'Qtd. pessoas': 'qtd_pessoas',
    'Logradouro': 'logradouro'
})

# Converter e limpar dados
df['data'] = pd.to_datetime(df['data'], errors='coerce')
df = df.dropna(subset=['data'])
df['qtd_pessoas'] = pd.to_numeric(df['qtd_pessoas'], errors='coerce')
df = df.dropna(subset=['qtd_pessoas'])

# Normalizar períodos
df['periodo_norm'] = df['periodo'].apply(normalizar_periodo)
df = df.dropna(subset=['periodo_norm'])

print(f"✓ Dados preparados")

# %% [markdown]
# # 7. Calcular Média Anterior (Intervalo Deslocado)

# %%
# Média anterior = mesmo intervalo, mas deslocado 1 dia para trás
# Exemplo: Se atual é 19/10 a 22/10, anterior é 18/10 a 21/10

data_inicio_anterior = data_inicio - timedelta(days=1)
data_fim_anterior = data_fim - timedelta(days=1)

df_anterior = df[(df['data'] >= data_inicio_anterior) & (df['data'] <= data_fim_anterior)].copy()

if len(df_anterior) > 0:
    # Criar contagens do período anterior
    df_anterior['data_str'] = df_anterior['data'].dt.strftime('%d/%m/%Y')
    contagens_anterior = df_anterior.groupby(['logradouro', 'periodo_norm', 'data_str'])['qtd_pessoas'].sum().to_dict()
    
    # Gerar lista de dias do período anterior
    dias_anteriores = gerar_lista_dias(data_inicio_anterior, data_fim_anterior)
    dias_validos_anterior = dias_anteriores[1:]  # Para manhã/tarde/madrugada
    dias_noite_anterior = dias_anteriores[:-1]   # Noite fecha no dia anterior
    
    # Obter logradouros únicos do período anterior com ordenação
    df_logradouros_unicos_anterior = df_anterior[['logradouro', 'tipo_logradouro', 'nome_logradouro', 'numero_logradouro']].drop_duplicates()
    df_logradouros_ordenados_anterior = ordenar_logradouros_df(df_logradouros_unicos_anterior)
    logradouros_anterior = df_logradouros_ordenados_anterior['logradouro'].tolist()
    
    # Construir matriz de totais por coluna (mesmo cálculo da média atual)
    totais_por_coluna_anterior = []
    
    for periodo in ['madrugada', 'manhã', 'tarde', 'noite']:
        dias_ref = dias_noite_anterior if periodo == 'noite' else dias_validos_anterior
        
        for dia in dias_ref:
            dia_str = dia.strftime('%d/%m/%Y')
            soma_coluna = 0
            
            for logradouro in logradouros_anterior:
                chave = (logradouro, periodo, dia_str)
                valor = contagens_anterior.get(chave, 0)
                soma_coluna += valor
            
            if soma_coluna > 0:
                totais_por_coluna_anterior.append(soma_coluna)
    
    # Calcular média anterior (mesma lógica da média atual)
    if totais_por_coluna_anterior:
        media_anterior = round(sum(totais_por_coluna_anterior) / len(totais_por_coluna_anterior))
    else:
        media_anterior = 0
    
    print(f"✓ Média anterior calculada: {media_anterior:.0f} pessoas/dia")
    print(f"  • Intervalo anterior: {data_inicio_anterior.strftime('%d/%m/%Y')} a {data_fim_anterior.strftime('%d/%m/%Y')}")
    print(f"  • Total de colunas calculadas: {len(totais_por_coluna_anterior)}")
else:
    media_anterior = 0
    print(f"⚠️  Sem dados para o intervalo anterior ({data_inicio_anterior.strftime('%d/%m/%Y')} a {data_fim_anterior.strftime('%d/%m/%Y')})")
    print(f"  • Média anterior = 0")

# Filtrar pelo período solicitado (atual)
df_periodo = df[(df['data'] >= data_inicio) & (df['data'] <= data_fim)].copy()

print(f"✓ Dados do período atual: {len(df_periodo):,} registros")

# %% [markdown]
# # 8. Gerar Lista de Dias

# %%
dias_lista = gerar_lista_dias(data_inicio, data_fim)
dias_validos = dias_lista[1:]  # Para manhã/tarde/madrugada (a partir do 2º dia)
dias_noite = dias_lista[:-1]   # Noite "fecha" no dia anterior (até penúltimo dia)

print(f"\n📅 Estrutura dos dias:")
print(f"  • Total de dias: {len(dias_lista)}")
print(f"  • Dias válidos (manhã/tarde/madrugada): {len(dias_validos)}")
print(f"  • Dias noite: {len(dias_noite)}")

# %% [markdown]
# # 9. Construir Matriz de Contagens

# %%
print(f"\n🔄 Construindo matriz de contagens...")

# Criar coluna de data formatada
df_periodo['data_str'] = df_periodo['data'].dt.strftime('%d/%m/%Y')

# Agrupar e contar
contagens = df_periodo.groupby(['logradouro', 'periodo_norm', 'data_str'])['qtd_pessoas'].sum().to_dict()

# Obter logradouros únicos com seus dados de ordenação
df_logradouros_unicos = df_periodo[['logradouro', 'tipo_logradouro', 'nome_logradouro', 'numero_logradouro']].drop_duplicates()

# Ordenar usando a função personalizada
df_logradouros_ordenados = ordenar_logradouros_df(df_logradouros_unicos)

# Lista de logradouros na ordem correta
logradouros = df_logradouros_ordenados['logradouro'].tolist()

print(f"✓ {len(logradouros)} logradouros únicos identificados e ordenados")
print(f"  • Primeiros 5: {logradouros[:5]}")
print(f"  • Últimos 5: {logradouros[-5:]}")

# %% [markdown]
# # 10. Criar Cabeçalhos

# %%
periodos = ['madrugada', 'manhã', 'tarde', 'noite']
periodos_fmt = {
    'madrugada': 'Madrugada',
    'manhã': 'Manhã',
    'tarde': 'Tarde',
    'noite': 'Noite'
}

# Header 1: Título do relatório
primeiro_dia = dias_validos[0] if dias_validos else data_inicio
ultimo_dia = dias_validos[-1] if dias_validos else data_fim
header1 = [f"Contagem diária - Santa Cecília, Campos Elíseos e Santa Ifigênia - {primeiro_dia.strftime('%d/%m/%Y')} a {ultimo_dia.strftime('%d/%m/%Y')}"]

# Header 2: Períodos
header2 = ['Ordem', 'Período']

for periodo in periodos:
    dias_ref = dias_noite if periodo == 'noite' else dias_validos
    for _ in dias_ref:
        header2.append(periodos_fmt[periodo])

# Adicionar colunas de médias por período e contagem >10
header2.extend(['Média por período', '', '', '', ''])

# Header 3: Dias
header3 = ['', 'Logradouro' + ' ' * 20 + 'Data']

for periodo in periodos:
    dias_ref = dias_noite if periodo == 'noite' else dias_validos
    for dia in dias_ref:
        header3.append(dia.strftime('%d'))

# Adicionar cabeçalhos das novas colunas
header3.extend(['Madrugada', 'Manhã', 'Tarde', 'Noite', '>10'])

# Igualar larguras
colunas_totais = len(header2)
while len(header1) < colunas_totais:
    header1.append('')
while len(header3) < colunas_totais:
    header3.append('')

print(f"✓ Cabeçalhos criados: {colunas_totais} colunas")

# %% [markdown]
# # 11. Construir Matriz de Dados

# %%
print(f"\n🔄 Construindo matriz de dados...")

matriz = []
visiveis = []

for logradouro in logradouros:
    linha = ['', logradouro]
    soma_linha = 0
    alguma_acima_limiar = False
    
    # Dicionário para armazenar valores por período
    valores_por_periodo = {
        'madrugada': [],
        'manhã': [],
        'tarde': [],
        'noite': []
    }
    
    # Contar células > 10
    contador_acima_10 = 0
    
    for periodo in periodos:
        dias_ref = dias_noite if periodo == 'noite' else dias_validos
        
        for dia in dias_ref:
            dia_str = dia.strftime('%d/%m/%Y')
            chave = (logradouro, periodo, dia_str)
            valor = contagens.get(chave, 0)
            
            linha.append(valor if valor > 0 else '')
            soma_linha += valor
            
            if valor > LIMIAR:
                alguma_acima_limiar = True
                contador_acima_10 += 1
            
            # Guardar valores > 0 para calcular média por período
            if valor > 0:
                valores_por_periodo[periodo].append(valor)
    
    # Calcular médias por período (apenas valores > 0)
    media_madrugada = round(sum(valores_por_periodo['madrugada']) / len(valores_por_periodo['madrugada'])) if valores_por_periodo['madrugada'] else ''
    media_manha = round(sum(valores_por_periodo['manhã']) / len(valores_por_periodo['manhã'])) if valores_por_periodo['manhã'] else ''
    media_tarde = round(sum(valores_por_periodo['tarde']) / len(valores_por_periodo['tarde'])) if valores_por_periodo['tarde'] else ''
    media_noite = round(sum(valores_por_periodo['noite']) / len(valores_por_periodo['noite'])) if valores_por_periodo['noite'] else ''
    
    # Adicionar as 5 novas colunas
    linha.extend([
        media_madrugada,
        media_manha,
        media_tarde,
        media_noite,
        contador_acima_10 if contador_acima_10 > 0 else ''
    ])
    
    if soma_linha > 0:
        matriz.append(linha)
        visiveis.append(alguma_acima_limiar)

# Numerar apenas as linhas visíveis
ordem = 1
for i in range(len(matriz)):
    if visiveis[i]:
        matriz[i][0] = ordem
        ordem += 1
    else:
        matriz[i][0] = ''

print(f"✓ Matriz criada: {len(matriz)} logradouros")
print(f"  • Visíveis (com aglomeração >{LIMIAR}): {sum(visiveis)}")
print(f"  • Ocultos: {len(visiveis) - sum(visiveis)}")

# %% [markdown]
# # 12. Calcular Linha de Totais

# %%
total_row = [''] * colunas_totais
total_row[1] = 'TOTAL'

# Calcular totais das colunas de dados (ignora as 5 últimas que são as médias e contagem)
num_colunas_dados = colunas_totais - 5  # Descontar as 5 novas colunas

for col in range(2, num_colunas_dados):
    soma = sum(matriz[row][col] if isinstance(matriz[row][col], (int, float)) else 0 
               for row in range(len(matriz)))
    total_row[col] = soma if soma > 0 else ''

# Calcular médias por período para a linha TOTAL
total_valores_por_periodo = {
    'madrugada': [],
    'manhã': [],
    'tarde': [],
    'noite': []
}

# Índice das colunas de dados por período
col_idx = 2
for periodo in periodos:
    dias_ref = dias_noite if periodo == 'noite' else dias_validos
    for _ in dias_ref:
        valor = total_row[col_idx]
        if isinstance(valor, (int, float)) and valor > 0:
            total_valores_por_periodo[periodo].append(valor)
        col_idx += 1

# Calcular médias dos totais
media_total_madr = round(sum(total_valores_por_periodo['madrugada']) / len(total_valores_por_periodo['madrugada'])) if total_valores_por_periodo['madrugada'] else ''
media_total_manha = round(sum(total_valores_por_periodo['manhã']) / len(total_valores_por_periodo['manhã'])) if total_valores_por_periodo['manhã'] else ''
media_total_tarde = round(sum(total_valores_por_periodo['tarde']) / len(total_valores_por_periodo['tarde'])) if total_valores_por_periodo['tarde'] else ''
media_total_noite = round(sum(total_valores_por_periodo['noite']) / len(total_valores_por_periodo['noite'])) if total_valores_por_periodo['noite'] else ''

# Contar total de células > 10
total_acima_10 = sum(1 for col in range(2, num_colunas_dados) 
                     if isinstance(total_row[col], (int, float)) and total_row[col] > LIMIAR)

# Adicionar as 5 colunas no final da linha TOTAL
total_row[num_colunas_dados] = media_total_madr
total_row[num_colunas_dados + 1] = media_total_manha
total_row[num_colunas_dados + 2] = media_total_tarde
total_row[num_colunas_dados + 3] = media_total_noite
total_row[num_colunas_dados + 4] = ''  # Coluna >10 não tem total

# Calcular média atual (para o texto de análise)
valores_somados = [v for v in total_row[2:num_colunas_dados] if isinstance(v, (int, float)) and v > 0]
media_atual = round(sum(valores_somados) / len(valores_somados)) if valores_somados else 0

print(f"\n📊 Médias calculadas:")
print(f"  • Média atual: {media_atual} pessoas/dia")
print(f"  • Média anterior: {media_anterior} pessoas/dia")
print(f"\n🔍 Debug linha TOTAL:")
print(f"  • Total Madrugada (col {num_colunas_dados}): {total_row[num_colunas_dados]}")
print(f"  • Total Manhã (col {num_colunas_dados + 1}): {total_row[num_colunas_dados + 1]}")
print(f"  • Total Tarde (col {num_colunas_dados + 2}): {total_row[num_colunas_dados + 2]}")
print(f"  • Total Noite (col {num_colunas_dados + 3}): {total_row[num_colunas_dados + 3]}")
print(f"  • Total >10 (col {num_colunas_dados + 4}): {total_row[num_colunas_dados + 4]}")

# %% [markdown]
# # 13. Gerar Texto de Análise

# %%
def somar_periodo_no_dia(periodo, dia_str):
    """Calcula totais de um período em um dia específico"""
    total = 0
    enderecos = 0
    soma_aglom = 0
    
    for logradouro in logradouros:
        chave = (logradouro, periodo, dia_str)
        valor = contagens.get(chave, 0)
        total += valor
        
        if valor > LIMIAR:
            enderecos += 1
            soma_aglom += valor
    
    return {'total': total, 'enderecos': enderecos, 'soma_aglom': soma_aglom}

# Último dia de cada tipo
ultimo_dia_val = dias_validos[-1] if dias_validos else data_fim
ultimo_dia_noite = dias_noite[-1] if dias_noite else data_fim

# Estatísticas do último dia
madr = somar_periodo_no_dia('madrugada', ultimo_dia_val.strftime('%d/%m/%Y'))
manha = somar_periodo_no_dia('manhã', ultimo_dia_val.strftime('%d/%m/%Y'))
tarde = somar_periodo_no_dia('tarde', ultimo_dia_val.strftime('%d/%m/%Y'))
noite = somar_periodo_no_dia('noite', ultimo_dia_noite.strftime('%d/%m/%Y'))

# Calcular os 5 logradouros com maior frequência nos últimos 3 dias
# Pegar os últimos 3 dias do período
if len(dias_validos) >= 3:
    ultimos_3_dias = dias_validos[-3:]
else:
    ultimos_3_dias = dias_validos

# Somar total por logradouro nos últimos 3 dias
soma_por_logradouro = {}
for logradouro in logradouros:
    total = 0
    for dia in ultimos_3_dias:
        dia_str = dia.strftime('%d/%m/%Y')
        for periodo in periodos:
            chave = (logradouro, periodo, dia_str)
            valor = contagens.get(chave, 0)
            total += valor
    
    # Adicionar também a noite dos dias correspondentes
    for dia in ultimos_3_dias:
        dia_str = dia.strftime('%d/%m/%Y')
        chave = (logradouro, 'noite', dia_str)
        valor = contagens.get(chave, 0)
        total += valor
    
    if total > 0:
        soma_por_logradouro[logradouro] = total

# Ordenar e pegar os 5 maiores
top_5_logradouros = sorted(soma_por_logradouro.items(), key=lambda x: x[1], reverse=True)[:5]

# Formatar texto dos top 5
if len(top_5_logradouros) >= 5:
    top_5_texto = "; ".join([f"{log}" for log, _ in top_5_logradouros[:-1]])
    top_5_texto += f" e {top_5_logradouros[-1][0]}"
else:
    # Se tiver menos de 5, ajustar formatação
    top_5_texto = "; ".join([f"{log}" for log, _ in top_5_logradouros])

# Calcular variação
if media_anterior > 0:
    variacao = round(((media_atual - media_anterior) / media_anterior) * 100, 1)
else:
    variacao = 0

tipo_variacao = "um aumento" if variacao > 0 else ("uma diminuição" if variacao < 0 else "estabilidade")

# Referência temporal
hoje = datetime.now()
dia_semana = hoje.weekday()
ref_texto = "sexta-feira" if dia_semana == 0 else "ontem"

# Texto final
texto_analise = (
    f"Na região de Santa Cecília, Campos Elíseos e Santa Ifigênia, em {ultimo_dia_val.strftime('%d/%m/%Y')} "
    f"foram localizadas {madr['total']} pessoas de madrugada (05h), {manha['total']} de manhã (10h), "
    f"{tarde['total']} à tarde (15h) e {noite['total']} à noite (20h) do dia {ultimo_dia_noite.strftime('%d')}. "
    f"Os 5 logradouros com maior frequência nos últimos 3 dias são: {top_5_texto}. "
    f"Com mais de 10 pessoas, foram {madr['enderecos']} endereços de madrugada, {manha['enderecos']} de manhã, "
    f"{tarde['enderecos']} à tarde e {noite['enderecos']} à noite, "
    f"somando respectivamente {madr['soma_aglom']}, {manha['soma_aglom']}, {tarde['soma_aglom']} e {noite['soma_aglom']}. "
    f"A média atual é de {media_atual} pessoas por dia — {tipo_variacao} de {abs(variacao)}% "
    f"em relação à contagem enviada {ref_texto}."
)

print(f"\n📝 Texto de análise gerado")

# %% [markdown]
# # 14. Criar Rodapé

# %%
hoje_formatado = hoje.strftime('%d/%m/%Y')

rodape = [
    ['Nota: As ruas sem aglomeração (>10) no período solicitado estão ocultas, mas constam na planilha.'],
    ['Fonte: SMS/Redenção na Rua'],
    [f'Elaborado por: SGM/SEPE, em {hoje_formatado}']
]

# Normalizar largura do rodapé
rodape_norm = []
for linha in rodape:
    linha_completa = linha + [''] * (colunas_totais - len(linha))
    rodape_norm.append(linha_completa)

print(f"✓ Rodapé criado")

# %% [markdown]
# # 15. Montar Saída Completa

# %%
saida = [
    header1,
    header2,
    header3,
    *matriz,
    total_row,
    *rodape_norm
]

print(f"\n✓ Saída montada: {len(saida)} linhas × {colunas_totais} colunas")
print(f"  • Colunas de dados: 3 até {num_colunas_dados - 1}")
print(f"  • Coluna Madrugada: {num_colunas_dados}")
print(f"  • Coluna Manhã: {num_colunas_dados + 1}")
print(f"  • Coluna Tarde: {num_colunas_dados + 2}")
print(f"  • Coluna Noite: {num_colunas_dados + 3}")
print(f"  • Coluna >10: {num_colunas_dados + 4}")
print(f"  • Total de colunas: {colunas_totais}")

# %% [markdown]
# # 16. Exportar para Excel

# %%
nome_arquivo_saida = f"relatorio_diario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
caminho_saida = DOCS_DIR / nome_arquivo_saida

print(f"\n💾 Exportando para Excel...")
print("-" * 80)

# Criar DataFrame
df_saida = pd.DataFrame(saida)

# Salvar sem formatação primeiro
df_saida.to_excel(caminho_saida, index=False, header=False, engine='openpyxl')

print(f"✓ Arquivo base criado: {nome_arquivo_saida}")

# %% [markdown]
# # 17. Aplicar Formatação

# %%
print(f"\n🎨 Aplicando formatação...")

# Carregar workbook
wb = load_workbook(caminho_saida)
ws = wb.active

# Estilos
fonte_bold = Font(bold=True)
fonte_italic = Font(italic=True, size=10)
fill_cinza = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
fill_azul = PatternFill(start_color='B7E1FA', end_color='B7E1FA', fill_type='solid')
border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Header (linhas 1-3)
for row in range(1, 4):
    for col in range(1, colunas_totais + 1):
        cell = ws.cell(row, col)
        cell.font = fonte_bold
        cell.fill = fill_cinza
        cell.border = border_thin

# Título (linha 1) - mesclar
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=colunas_totais)
ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')

# Mesclar "Ordem" e "Período" (linha 2-3, coluna 1)
ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
ws.cell(2, 1).alignment = Alignment(horizontal='center', vertical='center')

# Mesclar períodos na linha 2
col_inicio = 3
for periodo in periodos:
    qtd_dias = len(dias_noite) if periodo == 'noite' else len(dias_validos)
    if qtd_dias > 0:
        ws.merge_cells(start_row=2, start_column=col_inicio, end_row=2, end_column=col_inicio + qtd_dias - 1)
        ws.cell(2, col_inicio).alignment = Alignment(horizontal='center', vertical='center')
    col_inicio += qtd_dias

# Mesclar "Média por período" (4 colunas)
col_medias = colunas_totais - 4  # Últimas 5 colunas menos a última (>10)
ws.merge_cells(start_row=2, start_column=col_medias, end_row=2, end_column=col_medias + 3)
ws.cell(2, col_medias).value = "Média por período"
ws.cell(2, col_medias).alignment = Alignment(horizontal='center', vertical='center')

# Centralizar dias (linha 3)
for col in range(3, colunas_totais + 1):
    ws.cell(3, col).alignment = Alignment(horizontal='center')

# Definir índices das colunas das novas features (para usar na formatação)
col_madrugada = colunas_totais - 4
col_manha = colunas_totais - 3  
col_tarde = colunas_totais - 2
col_noite = colunas_totais - 1
col_maior10 = colunas_totais

# Mesclar ">10" (linha 2-3, última coluna) - DEPOIS de definir col_maior10
ws.merge_cells(start_row=2, start_column=col_maior10, end_row=3, end_column=col_maior10)
ws.cell(2, col_maior10).value = ">10"
ws.cell(2, col_maior10).alignment = Alignment(horizontal='center', vertical='center')

# Dados - centralizar e aplicar formatação condicional
primeira_linha_dados = 4

for row in range(primeira_linha_dados, primeira_linha_dados + len(matriz)):
    # Coluna ordem
    ws.cell(row, 1).alignment = Alignment(horizontal='center')
    
    # Colunas de dados diários (coluna 3 até antes das médias)
    for col in range(3, col_madrugada):
        cell = ws.cell(row, col)
        cell.alignment = Alignment(horizontal='center')
        cell.border = border_thin
        
        # Formatação condicional (> LIMIAR = azul claro)
        if isinstance(cell.value, (int, float)) and cell.value > LIMIAR:
            cell.fill = fill_azul
    
    # Coluna Madrugada
    cell = ws.cell(row, col_madrugada)
    cell.alignment = Alignment(horizontal='center')
    cell.border = border_thin
    if isinstance(cell.value, (int, float)) and cell.value > LIMIAR:
        cell.fill = fill_azul
    
    # Coluna Manhã
    cell = ws.cell(row, col_manha)
    cell.alignment = Alignment(horizontal='center')
    cell.border = border_thin
    if isinstance(cell.value, (int, float)) and cell.value > LIMIAR:
        cell.fill = fill_azul
    
    # Coluna Tarde
    cell = ws.cell(row, col_tarde)
    cell.alignment = Alignment(horizontal='center')
    cell.border = border_thin
    if isinstance(cell.value, (int, float)) and cell.value > LIMIAR:
        cell.fill = fill_azul
    
    # Coluna Noite
    cell = ws.cell(row, col_noite)
    cell.alignment = Alignment(horizontal='center')
    cell.border = border_thin
    if isinstance(cell.value, (int, float)) and cell.value > LIMIAR:
        cell.fill = fill_azul
    
    # Coluna >10
    # COM borda e centralização, SEM azul
    cell = ws.cell(row, col_maior10)
    cell.alignment = Alignment(horizontal='center')
    cell.border = border_thin
    # Não aplica fill azul

# Linha TOTAL
linha_total = primeira_linha_dados + len(matriz)

# Formatar todas as colunas EXCETO a última (>10)
for col in range(1, col_maior10):  
    cell = ws.cell(linha_total, col)
    cell.font = fonte_bold
    cell.fill = fill_cinza
    cell.alignment = Alignment(horizontal='center')
    cell.border = border_thin

# Última coluna do TOTAL (>10) - SEM formatação, campo vazio
cell = ws.cell(linha_total, col_maior10)
cell.value = ''
# Não aplica nenhuma formatação

# Rodapé (depois da linha TOTAL)
linha_rodape_inicio = linha_total + 1
for row in range(linha_rodape_inicio, linha_rodape_inicio + 3):
    for col in range(1, colunas_totais + 1):
        cell = ws.cell(row, col)
        cell.font = fonte_italic
        cell.alignment = Alignment(horizontal='left')

# Linha MÉDIA (DEPOIS do rodapé, não antes)
linha_media = linha_rodape_inicio + 4  # 3 linhas de rodapé + 1 linha vazia
ws.cell(linha_media, 1).value = ''
ws.cell(linha_media, 2).value = 'Média:'
ws.cell(linha_media, 2).font = fonte_bold
ws.cell(linha_media, 2).alignment = Alignment(horizontal='left')

ws.cell(linha_media, 3).value = media_atual
ws.cell(linha_media, 3).font = fonte_bold
ws.cell(linha_media, 3).alignment = Alignment(horizontal='center')

# Ajustar larguras das colunas
ws.column_dimensions['A'].width = 8   # Ordem
ws.column_dimensions['B'].width = 45  # Logradouro (aumentado para não cortar)

# Colunas de dados diários
for col in range(3, col_madrugada):
    ws.column_dimensions[get_column_letter(col)].width = 6

# Colunas de médias e >10 (largura maior para não cortar)
for col in range(col_madrugada, col_maior10 + 1):
    ws.column_dimensions[get_column_letter(col)].width = 12

# Ocultar linhas sem aglomeração
for i, visivel in enumerate(visiveis):
    if not visivel:
        ws.row_dimensions[primeira_linha_dados + i].hidden = True

# Salvar
wb.save(caminho_saida)

print(f"✓ Formatação aplicada")
print(f"✓ {len(visiveis) - sum(visiveis)} linhas ocultas (sem aglomeração >{LIMIAR})")
print(f"✓ Linha de MÉDIA adicionada (após rodapé) com valor: {media_atual}")

# %% [markdown]
# # 18. Exportar Texto de Análise para TXT

# %%
print(f"\n📝 Exportando texto de análise...")

# Nome do arquivo TXT (mesmo nome base do Excel)
nome_base = nome_arquivo_saida.replace('.xlsx', '')
nome_txt = f"{nome_base}_analise.txt"
caminho_txt = DOCS_DIR / nome_txt

# Criar conteúdo do TXT
conteudo_txt = f"""================================================================================
TEXTO DE ANÁLISE - RELATÓRIO DIÁRIO
================================================================================

Período do Relatório: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}
Gerado em: {hoje.strftime('%d/%m/%Y às %H:%M:%S')}

================================================================================
ANÁLISE
================================================================================

{texto_analise}

================================================================================
ESTATÍSTICAS
================================================================================

Média Atual:    {media_atual} pessoas/dia (intervalo {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')})
Média Anterior: {media_anterior:.0f} pessoas/dia (intervalo {data_inicio_anterior.strftime('%d/%m/%Y')} a {data_fim_anterior.strftime('%d/%m/%Y')})
Variação:       {variacao:+.1f}%

================================================================================
ÚLTIMO DIA ANALISADO - {ultimo_dia_val.strftime('%d/%m/%Y')}
================================================================================

Madrugada (05h):
  • Total de pessoas: {madr['total']}
  • Endereços com >10 pessoas: {madr['enderecos']}
  • Soma nas aglomerações: {madr['soma_aglom']}

Manhã (10h):
  • Total de pessoas: {manha['total']}
  • Endereços com >10 pessoas: {manha['enderecos']}
  • Soma nas aglomerações: {manha['soma_aglom']}

Tarde (15h):
  • Total de pessoas: {tarde['total']}
  • Endereços com >10 pessoas: {tarde['enderecos']}
  • Soma nas aglomerações: {tarde['soma_aglom']}

Noite (20h) do dia {ultimo_dia_noite.strftime('%d/%m/%Y')}:
  • Total de pessoas: {noite['total']}
  • Endereços com >10 pessoas: {noite['enderecos']}
  • Soma nas aglomerações: {noite['soma_aglom']}

================================================================================
"""

# Salvar arquivo TXT
with open(caminho_txt, 'w', encoding='utf-8') as f:
    f.write(conteudo_txt)

print(f"✓ Texto exportado: {nome_txt}")

# %% [markdown]
# # 19. Resumo Executivo

# %%
print(f"\n" + "=" * 80)
print("RESUMO EXECUTIVO")
print("=" * 80)

print(f"\n📊 Relatório Gerado:")
print(f"  • Período atual: {data_inicio.strftime('%d/%m/%Y')} até {data_fim.strftime('%d/%m/%Y')}")
print(f"  • Período anterior: {data_inicio_anterior.strftime('%d/%m/%Y')} até {data_fim_anterior.strftime('%d/%m/%Y')}")
print(f"  • Total de dias: {len(dias_lista)}")
print(f"  • Logradouros analisados: {len(logradouros)}")
print(f"  • Logradouros visíveis (>{LIMIAR}): {sum(visiveis)}")
print(f"  • Logradouros ocultos: {len(visiveis) - sum(visiveis)}")

print(f"\n📈 Estatísticas:")
print(f"  • Média atual: {media_atual} pessoas/dia")
print(f"  • Média anterior: {media_anterior:.0f} pessoas/dia")
print(f"  • Variação: {variacao:+.1f}%")
print(f"  • Tipo de variação: {tipo_variacao}")

print(f"\n📁 Arquivos gerados:")
print(f"  • Planilha: {nome_arquivo_saida}")
print(f"  • Texto: {nome_txt}")
print(f"  • Localização: {DOCS_DIR}")

print(f"\n📋 Último dia analisado ({ultimo_dia_val.strftime('%d/%m/%Y')}):")
print(f"  • Madrugada: {madr['total']} pessoas ({madr['enderecos']} endereços >{LIMIAR})")
print(f"  • Manhã: {manha['total']} pessoas ({manha['enderecos']} endereços >{LIMIAR})")
print(f"  • Tarde: {tarde['total']} pessoas ({tarde['enderecos']} endereços >{LIMIAR})")
print(f"  • Noite: {noite['total']} pessoas ({noite['enderecos']} endereços >{LIMIAR})")

print(f"\n✅ Relatório consolidado gerado com sucesso!")
print("=" * 80)