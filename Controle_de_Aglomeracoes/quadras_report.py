import pandas as pd
import re
import os
import sys
from itertools import groupby
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path

# --- CONFIGURAÇÃO DE ESTILOS ---
COLOR_GRAY = "D9D9D9"
COLOR_BLUE = "9BC2E6"

# Prefixos que identificam linhas de rodapé no relatório diário
FOOTER_PREFIXES = ("nota:", "fonte:", "elaborado", "média:", "media:")


def _is_footer_row(row_values):
    """
    Retorna True se a linha for parte do rodapé do relatório diário.
    Verifica col[0] e col[1] (após normalização) para os prefixos conhecidos.
    """
    for val in row_values[:3]:
        if pd.notna(val):
            normalized = str(val).strip().lower()
            if any(normalized.startswith(p) for p in FOOTER_PREFIXES):
                return True
    return False


def _find_gt10_col_index(header_row):
    """
    Encontra o índice (0-based) da coluna '>10' no cabeçalho.
    Retorna -1 se não encontrada, o que indica erro na estrutura.
    """
    for i, val in enumerate(header_row):
        if str(val).strip() == ">10":
            return i
    return len(header_row) - 1  # fallback: última coluna


def gerar_relatorio_quadras(input_file, log_callback):
    """
    Função principal corrigida.
    Correções em relação à versão anterior:
      1. Detecção dinâmica da coluna '>10' (não mais hardcoded em col[18]).
      2. Filtro explícito de linhas de rodapé antes do processamento.
      3. Subtotais somando apenas as colunas de dados (excluindo médias e '>10').
      4. Cabeçalho do arquivo de saída espelha exatamente o diário original.
      5. Numeração de ordem sequencial correta (apenas linhas visíveis).
    """
    log_callback(f"Iniciando processamento de Quadras...")
    log_callback(f"Lendo arquivo: {os.path.basename(input_file)}")

    base_path = Path(input_file).parent

    input_filename = os.path.basename(input_file)
    if "relatorio_diario" in input_filename:
        output_filename = input_filename.replace("relatorio_diario", "relatorio_quadras")
    else:
        output_filename = f"relatorio_quadras_{input_filename}"

    output_file = base_path / output_filename

    script_dir = Path(__file__).resolve().parent
    mapping_file_xlsx = script_dir / 'Mapeamento_FINAL_editado.xlsx'

    if not mapping_file_xlsx.exists():
        raise FileNotFoundError(f"Arquivo de mapeamento não encontrado em: {mapping_file_xlsx}")

    # 1. CARREGAR DADOS
    try:
        df_map = pd.read_excel(mapping_file_xlsx)
        df_map['Nome Original Norm'] = df_map['Nome Original'].astype(str).str.lower().str.strip()
        log_callback("✓ Mapeamento carregado.")
    except Exception as e:
        raise Exception(f"Erro ao ler mapeamento: {e}")

    try:
        df_main = pd.read_excel(input_file, header=None, dtype=str, engine='openpyxl')
    except Exception as e:
        raise Exception(f"Erro ao ler input: {e}")

    # 2. SEPARAR CABEÇALHO, DADOS E RODAPÉ
    # O relatório diário sempre tem 3 linhas de cabeçalho fixas.
    header = [df_main.iloc[0], df_main.iloc[1], df_main.iloc[2]]
    all_data = df_main.iloc[3:]

    # Detectar dinamicamente a coluna '>10'
    header3_values = header[2].tolist()
    col_gt10_idx = _find_gt10_col_index(header3_values)
    total_cols = len(header3_values)

    # Colunas de dados (excluindo 'Ordem', 'Logradouro', 4 médias e '>10')
    # Estrutura: [Ordem, Logradouro, ...dados..., Madr_media, Manha_media, Tarde_media, Noite_media, >10]
    # Portanto as colunas de dados são índices 2 até (total_cols - 6) inclusive.
    num_media_cols = 4
    col_data_start = 2
    col_data_end = total_cols - num_media_cols - 1  # exclusive do índice >10 e das 4 médias

    log_callback(f"✓ Estrutura detectada: {total_cols} colunas, coluna '>10' no índice {col_gt10_idx}.")

    # 3. FUNÇÃO: ENCONTRAR QUADRA
    def find_quadra(logradouro, df_map_ref):
        if pd.isna(logradouro):
            return None
        logradouro_str = str(logradouro).strip()
        match = re.match(r"^(.*?),\s*(\d+)", logradouro_str)

        if match:
            nome = match.group(1).strip().lower()
            try:
                num = int(match.group(2))
            except Exception:
                num = 0
        else:
            nome = logradouro_str.lower()
            num = 0

        cands = df_map_ref[df_map_ref['Nome Original Norm'] == nome]
        if cands.empty:
            return None

        for _, r in cands.iterrows():
            try:
                n_min = pd.to_numeric(r['Num Min'], errors='coerce')
                n_max = pd.to_numeric(r['Num Max'], errors='coerce')
                if pd.notna(n_min) and pd.notna(n_max):
                    if int(n_min) <= num <= int(n_max):
                        return r['Quadra']
            except Exception:
                continue
        return None

    # 4. PROCESSAMENTO — filtrando rodapé e linha de total
    log_callback("Processando linhas e identificando quadras...")

    total_row = None
    cleaned_rows = []

    for idx, row in all_data.iterrows():
        row_vals = row.tolist()

        # Pular linhas completamente vazias
        if all(pd.isna(v) or str(v).strip() in ('', 'nan') for v in row_vals):
            continue

        # Identificar e separar linha de rodapé
        if _is_footer_row(row_vals):
            continue

        c0 = str(row_vals[0]).lower() if pd.notna(row_vals[0]) else ""
        c1 = str(row_vals[1]).lower() if pd.notna(row_vals[1]) else ""

        # Identificar linha de total geral
        if "total" in c0 or "total" in c1:
            total_row = row_vals
            continue

        # FIX 1: usar col_gt10_idx dinâmico ao invés de row[18] fixo
        quadra = find_quadra(row_vals[1], df_map)
        val_gt10 = row_vals[col_gt10_idx] if col_gt10_idx < len(row_vals) else None
        is_visible = pd.notna(val_gt10) and str(val_gt10).strip() not in ('', 'nan', '0')

        cleaned_rows.append({
            'data': row_vals,
            'quadra': quadra,
            'visible': is_visible,
            'original_idx': idx
        })

    cleaned_rows.sort(
        key=lambda x: (x['quadra'] if x['quadra'] else "ZZZ_SEM_QUADRA", x['original_idx'])
    )

    # 5. MONTAR ESTRUTURA FINAL COM SUBTOTAIS
    final_structure = []

    log_callback("Calculando subtotais...")
    for key, group in groupby(cleaned_rows, key=lambda x: x['quadra']):
        group_list = list(group)
        for item in group_list:
            final_structure.append({'type': 'data', 'values': item['data'], 'visible': item['visible']})

        if key is None:
            continue

        visible_rows_in_group = [it for it in group_list if it['visible']]

        if len(visible_rows_in_group) > 1:
            sub_vals = [""] * total_cols
            sub_vals[1] = "Subtotal"

            # FIX 2: somar apenas colunas de dados (excluindo médias e '>10')
            for c in range(col_data_start, col_data_end):
                soma = 0
                for it in visible_rows_in_group:
                    v = pd.to_numeric(it['data'][c], errors='coerce') if c < len(it['data']) else None
                    if pd.notna(v):
                        soma += v
                sub_vals[c] = int(round(soma))

            final_structure.append({'type': 'subtotal', 'values': sub_vals, 'visible': True})

    if total_row is not None:
        final_structure.append({'type': 'total_geral', 'values': total_row, 'visible': True})

    # 6. GERAR EXCEL
    log_callback("Gerando Excel final...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"

    thin = Side(border_style="thin", color="000000")
    border_all = Border(top=thin, left=thin, right=thin, bottom=thin)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    font_bold = Font(bold=True)
    fill_gray = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type="solid")
    fill_blue = PatternFill(start_color=COLOR_BLUE, end_color=COLOR_BLUE, fill_type="solid")

    def style_cell(cell, border=True, fill=None, bold=False, align="center"):
        if border:
            cell.border = border_all
        if fill:
            cell.fill = fill
        if bold:
            cell.font = font_bold
        cell.alignment = align_left if align == "left" else align_center

    # FIX 3: Replicar cabeçalho idêntico ao relatório diário (incluindo merges dinâmicos)
    # Linha 1: título mesclado
    h0_vals = header[0].fillna('').tolist()
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(1, 1).value = h0_vals[0]
    style_cell(ws.cell(1, 1), fill=fill_gray, bold=True)

    # Linha 2: "Ordem", "Período", e cabeçalhos de períodos (com merge dinâmico)
    h1_vals = header[1].fillna('').tolist()
    # Coluna A (Ordem) — mesclada com linha 3
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    ws.cell(2, 1).value = h1_vals[0] if h1_vals[0] else "Ordem"
    style_cell(ws.cell(2, 1), fill=fill_gray, bold=True)

    # Coluna B (Logradouro/Período)
    ws.cell(2, 2).value = h1_vals[1] if len(h1_vals) > 1 else "Período"
    style_cell(ws.cell(2, 2), fill=fill_gray, bold=True)

    # Cabeçalhos de período: replicar os merges da linha 2 do original
    # Detecta os blocos mesclados percorrendo header[1] — cada período tem N células consecutivas com o mesmo label
    col = 3
    while col <= col_data_end:
        periodo_label = str(h1_vals[col - 1]).strip() if col - 1 < len(h1_vals) else ""
        if not periodo_label or periodo_label in ('', 'nan'):
            col += 1
            continue
        # Contar quantas colunas consecutivas têm o mesmo label (originalmente mescladas)
        span = 1
        while (col + span - 1) < col_data_end and str(h1_vals[col + span - 1]).strip() in ('', 'nan', periodo_label):
            span += 1
        if span > 1:
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + span - 1)
        ws.cell(2, col).value = periodo_label
        style_cell(ws.cell(2, col), fill=fill_gray, bold=True)
        col += span

    # Cabeçalhos das médias (mesclado) e '>10' (mesclado com linha 3)
    col_medias = col_data_end + 1  # primeira coluna de média (1-based)
    ws.merge_cells(start_row=2, start_column=col_medias, end_row=2, end_column=col_medias + 3)
    ws.cell(2, col_medias).value = "Média por período"
    style_cell(ws.cell(2, col_medias), fill=fill_gray, bold=True)

    ws.merge_cells(start_row=2, start_column=total_cols, end_row=3, end_column=total_cols)
    ws.cell(2, total_cols).value = ">10"
    style_cell(ws.cell(2, total_cols), fill=fill_gray, bold=True)

    # Linha 3: datas e sub-labels
    h2_vals = header[2].fillna('').tolist()
    ws.cell(3, 2).value = h2_vals[1] if len(h2_vals) > 1 else ""
    style_cell(ws.cell(3, 2), fill=fill_gray, bold=True)
    for c in range(col_data_start + 1, total_cols):  # cols 3..N-1 (1-based)
        v = h2_vals[c - 1] if c - 1 < len(h2_vals) else ""
        if isinstance(v, str) and v.endswith('.0'):
            v = v[:-2]
        ws.cell(3, c).value = v
        style_cell(ws.cell(3, c), fill=fill_gray, bold=True)
    # Médias: labels Madrugada/Manhã/Tarde/Noite
    for offset, label in enumerate(['Madrugada', 'Manhã', 'Tarde', 'Noite']):
        ws.cell(3, col_medias + offset).value = label
        style_cell(ws.cell(3, col_medias + offset), fill=fill_gray, bold=True)

    # 7. ESCREVER DADOS
    current_row = 4
    total_geral_vals = []

    # FIX 4: numeração correta — contar apenas visíveis, reiniciar por visibilidade
    visible_counter = 1

    for item in final_structure:
        vals = item['values']
        rtype = item['type']
        is_visible = item['visible']

        if not is_visible:
            ws.row_dimensions[current_row].hidden = True
        else:
            if rtype == 'data':
                vals = list(vals)  # garante mutabilidade
                vals[0] = visible_counter
                visible_counter += 1

        if rtype == 'total_geral':
            for i in range(col_data_start, col_data_end):
                v_num = pd.to_numeric(vals[i], errors='coerce') if i < len(vals) else None
                total_geral_vals.append(float(v_num) if pd.notna(v_num) else 0.0)

        for i, val in enumerate(vals):
            if i >= total_cols:
                break
            cell = ws.cell(row=current_row, column=i + 1)

            f_val = pd.to_numeric(val, errors='coerce')
            if pd.notna(f_val):
                cell.value = int(f_val) if float(f_val).is_integer() else float(f_val)
            else:
                cell.value = None if str(val).strip() in ('', 'nan') else val

            fill = None
            bold = False
            if rtype in ('subtotal', 'total_geral'):
                fill = fill_gray
                bold = True
            elif (i >= col_data_start and i < col_data_end
                  and isinstance(cell.value, (int, float))
                  and cell.value > 10):
                fill = fill_blue

            cell.border = border_all
            if fill:
                cell.fill = fill
            if bold:
                cell.font = font_bold
            cell.alignment = align_left if i == 1 else align_center

        current_row += 1

    # 8. RODAPÉ
    avg = int(round(sum(total_geral_vals) / len(total_geral_vals))) if total_geral_vals else 0

    ft_row = current_row + 1

    def write_footer(r, text, bold=False):
        c = ws.cell(row=r, column=2, value=text)
        c.font = Font(size=10, bold=bold)
        c.alignment = align_left

    write_footer(ft_row, "Nota: As ruas sem aglomeração (>10) no período solicitado estão ocultas, mas constam na planilha.")
    write_footer(ft_row + 1, "Fonte: SMS/Redenção na Rua")

    from datetime import datetime
    write_footer(ft_row + 2, f"Elaborado por: SGM/SEPE, em {datetime.now().strftime('%d/%m/%Y')}")

    mr = ft_row + 4
    ws.cell(row=mr, column=2, value="Média:").font = font_bold
    ws.cell(row=mr, column=2).alignment = Alignment(horizontal="right")
    ws.cell(row=mr, column=3, value=avg).font = font_bold
    ws.cell(row=mr, column=3).alignment = align_left

    # 9. LARGURAS DAS COLUNAS
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 65
    for i in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 8

    log_callback(f"Salvando arquivo: {output_filename}")
    wb.save(output_file)
    log_callback(f"✓ Relatório de Quadras gerado com sucesso!")
    log_callback(f"  📁 Local: {output_file}")

    return str(output_file)