import pandas as pd
import re
import os
import sys
from itertools import groupby
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path

# --- CONFIGURAÇÃO DE ESTILOS ---
COLOR_GRAY = "D9D9D9"
COLOR_BLUE = "9BC2E6"

def gerar_relatorio_quadras(input_file, log_callback):
    """
    Função principal chamada pelo main_app.py
    """
    log_callback(f"Iniciando processamento de Quadras...")
    log_callback(f"Lendo arquivo: {os.path.basename(input_file)}")

    # Define caminhos
    base_path = Path(input_file).parent
    
    # Nome do arquivo de saída (mantém o timestamp do input)
    input_filename = os.path.basename(input_file)
    if "relatorio_diario" in input_filename:
        output_filename = input_filename.replace("relatorio_diario", "relatorio_quadras")
    else:
        output_filename = f"relatorio_quadras_{input_filename}"
    
    output_file = base_path / output_filename
    
    # Arquivo de mapeamento (deve estar na mesma pasta do script ou do executável)
    script_dir = Path(__file__).resolve().parent
    mapping_file_xlsx = script_dir / 'Mapeamento_FINAL_editado.xlsx'

    if not mapping_file_xlsx.exists():
        raise FileNotFoundError(f"Arquivo de mapeamento não encontrado em: {mapping_file_xlsx}")

    # 1. CARREGAR DADOS
    try:
        df_map = pd.read_excel(mapping_file_xlsx)
        df_map['Nome Original Norm'] = df_map['Nome Original'].astype(str).str.lower().str.strip()
        log_callback("✓ Mapeamento carregado.")
    except Exception as e:
        raise Exception(f"Erro ao ler mapeamento: {e}")

    try:
        # Tenta ler com header=None para pegar o cabeçalho "bruto" das 3 primeiras linhas
        df_main = pd.read_excel(input_file, header=None, dtype=str, engine='openpyxl')
    except Exception as e:
        raise Exception(f"Erro ao ler input: {e}")

    # 2. FUNÇÃO: ENCONTRAR QUADRA
    def find_quadra(logradouro, df_map_ref):
        if pd.isna(logradouro): return None
        
        logradouro_str = str(logradouro).strip()
        
        # Tenta extrair "Rua, Número"
        match = re.match(r"^(.*?),\s*(\d+)", logradouro_str)
        
        if match:
            nome = match.group(1).strip().lower()
            num = int(match.group(2))
        else:
            # Assume que é só o nome da rua e número = 0
            nome = logradouro_str.lower()
            num = 0
        
        # Busca no mapeamento
        cands = df_map_ref[df_map_ref['Nome Original Norm'] == nome]
        if cands.empty: return None
        
        # Verifica intervalo
        for _, r in cands.iterrows():
            try:
                if int(r['Num Min']) <= num <= int(r['Num Max']):
                    return r['Quadra']
            except: continue
        return None

    # 3. PROCESSAMENTO
    log_callback("Processando linhas e identificando quadras...")
    
    # As 3 primeiras linhas são cabeçalho
    header = [df_main.iloc[0], df_main.iloc[1], df_main.iloc[2]]
    data = df_main.iloc[3:]

    total_row = None
    cleaned_rows = []

    for idx, row in data.iterrows():
        c0 = str(row[0]).lower() if pd.notna(row[0]) else ""
        c1 = str(row[1]).lower() if pd.notna(row[1]) else ""
        
        if "total" in c0 or "total" in c1:
            total_row = row
        else:
            quadra = find_quadra(row[1], df_map)
            val_gt10 = row[18] # Coluna S (índice 18)
            # Regra de visibilidade: Coluna S não pode ser vazia/NaN
            is_visible = pd.notna(val_gt10) and str(val_gt10).strip() != ""
            
            cleaned_rows.append({
                'data': row.tolist(), 
                'quadra': quadra, 
                'visible': is_visible, 
                'original_idx': idx
            })

    # Ordena por Quadra
    cleaned_rows.sort(key=lambda x: (x['quadra'] if x['quadra'] else "ZZZ_SEM_QUADRA", x['original_idx']))

    final_structure = []

    # Loop de Agrupamento
    log_callback("Calculando subtotais...")
    for key, group in groupby(cleaned_rows, key=lambda x: x['quadra']):
        group_list = list(group)
        
        # Adiciona dados brutos à estrutura final
        for item in group_list:
            final_structure.append({'type':'data', 'values':item['data'], 'visible':item['visible']})
        
        if key is None: continue
        
        # Filtra apenas linhas VISÍVEIS para o cálculo do subtotal
        visible_rows_in_group = [it for it in group_list if it['visible']]
        
        # Subtotal apenas se > 1 linha visível
        if len(visible_rows_in_group) > 1:
            sub_vals = [""] * 19
            sub_vals[1] = f"Subtotal"
            
            # Soma colunas C (idx 2) a R (idx 17)
            for c in range(2, 18):
                soma = 0
                for it in visible_rows_in_group:
                    try:
                        v = float(it['data'][c])
                        if pd.notna(v): soma += v
                    except: pass
                sub_vals[c] = int(round(soma))
            
            final_structure.append({'type':'subtotal', 'values':sub_vals, 'visible':True})

    # Adiciona Total Geral ao final
    if total_row is not None:
        final_structure.append({'type':'total_geral', 'values':total_row.tolist(), 'visible':True})

    # 4. WRITER EXCEL
    log_callback("Gerando Excel final...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"

    # Definições de Estilo
    thin = Side(border_style="thin", color="000000")
    border_all = Border(top=thin, left=thin, right=thin, bottom=thin)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    font_bold = Font(bold=True)
    fill_gray = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type="solid")
    fill_blue = PatternFill(start_color=COLOR_BLUE, end_color=COLOR_BLUE, fill_type="solid")

    def apply_style(rng, border=True, fill=None, bold=False, align="center"):
        rows = ws[rng] if ':' in rng else [[ws[rng]]]
        for r in rows:
            for c in r:
                if border: c.border = border_all
                if fill: c.fill = fill
                if bold: c.font = font_bold
                c.alignment = align_left if align == "left" else align_center

    # Cabeçalhos (Reconstrução visual)
    ws.merge_cells("A1:S1"); ws["A1"] = header[0][0]; apply_style("A1:S1", fill=fill_gray, bold=True)
    
    h1 = header[1].fillna('').tolist()
    ws.merge_cells("A2:A3"); ws["A2"] = h1[0]; apply_style("A2:A3", fill=fill_gray, bold=True)
    ws["B2"] = h1[1]; apply_style("B2", fill=fill_gray, bold=True)
    
    ranges = [("C2:E2",2),("F2:H2",5),("I2:K2",8),("L2:N2",11),("O2:R2",14),("S2:S3",18)]
    for rng, idx in ranges:
        ws.merge_cells(rng); ws[rng.split(':')[0]] = h1[idx]; apply_style(rng, fill=fill_gray, bold=True)
    
    h2 = header[2].fillna('').tolist()
    ws["B3"] = h2[1]; apply_style("B3", fill=fill_gray, bold=True)
    for i in range(2, 18):
        l = get_column_letter(i+1)
        v = h2[i]
        if isinstance(v, str) and v.endswith('.0'): v = v[:-2]
        ws[f"{l}3"] = v
        apply_style(f"{l}3", fill=fill_gray, bold=True)

    # Escrita dos Dados
    current_row = 4
    total_geral_vals = []
    visible_counter = 1 

    for item in final_structure:
        vals = item['values']
        rtype = item['type']
        is_visible = item['visible']
        
        if not is_visible:
            ws.row_dimensions[current_row].hidden = True
        else:
            if rtype == 'data':
                vals[0] = visible_counter
                visible_counter += 1
        
        # Coleta para média do rodapé
        if rtype == 'total_geral':
            for i in range(2, 14):
                try: total_geral_vals.append(float(vals[i]))
                except: pass
                
        for i, val in enumerate(vals):
            cell = ws.cell(row=current_row, column=i+1)
            try:
                f = float(val)
                val = int(f) if f.is_integer() else f
            except: pass
            cell.value = val
            
            # Estilo condicional
            fill = None
            bold = False
            if rtype in ['subtotal', 'total_geral']:
                fill = fill_gray
                bold = True
            elif i != 0 and i != 18 and isinstance(val, (int, float)) and val > 10:
                fill = fill_blue
                
            cell.border = border_all
            if fill: cell.fill = fill
            if bold: cell.font = font_bold
            cell.alignment = align_left if i == 1 else align_center
        current_row += 1

    # Rodapé
    avg = 0
    if total_geral_vals: avg = int(round(sum(total_geral_vals)/len(total_geral_vals)))

    ft_row = current_row + 1
    def wft(r, t, b=False):
        c = ws.cell(row=r, column=2, value=t)
        c.font = Font(size=10, bold=b)
        c.alignment = align_left

    wft(ft_row, "Nota: As ruas sem aglomeração (>10) no período solicitado estão ocultas, mas constam na planilha.")
    wft(ft_row+1, "Fonte: SMS/Redenção na Rua")
    
    # Data atual para o rodapé
    from datetime import datetime
    wft(ft_row+2, f"Elaborado por: SGM/SEPE, em {datetime.now().strftime('%d/%m/%Y')}")
    
    mr = ft_row + 4
    ws.cell(row=mr, column=2, value="Média:").font = font_bold
    ws.cell(row=mr, column=2).alignment = Alignment(horizontal="right")
    ws.cell(row=mr, column=3, value=avg).font = font_bold
    ws.cell(row=mr, column=3).alignment = align_left

    # Larguras das colunas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 65
    for i in range(3, 20): ws.column_dimensions[get_column_letter(i)].width = 8

    log_callback(f"Salvando arquivo: {output_filename}")
    wb.save(output_file)
    
    return str(output_file)