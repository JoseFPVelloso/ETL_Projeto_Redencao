# logic_report.py
# (Baseado em 05_processed_relatorio_diario_contagem_centro.py)
# VERS√ÉO ATUALIZADA: agora importa logic_text_generator e ordena varia√ß√µes extremas

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path
import warnings
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re
import traceback

# IMPORTA O NOVO M√ìDULO DE GERA√á√ÉO DE TEXTO
import logic_text_generator

warnings.filterwarnings('ignore')

# --- Fun√ß√µes Utilit√°rias (sem altera√ß√£o) ---

def normalizar_periodo(periodo_str):
    if not periodo_str: return None
    s = str(periodo_str).lower()
    if 'madrug' in s: return 'madrugada'
    if 'manh' in s or 'manha' in s: return 'manh√£'
    if 'tarde' in s: return 'tarde'
    if 'noite' in s: return 'noite'
    return None

def gerar_lista_dias(data_inicio, data_fim):
    dias = []
    data_atual = data_inicio
    while data_atual <= data_fim:
        dias.append(data_atual)
        data_atual += timedelta(days=1)
    return dias

def extrair_numero_logradouro(numero_str):
    if pd.isna(numero_str) or numero_str == '':
        return 999999
    try:
        digitos = re.search(r'\d+', str(numero_str))
        if digitos:
            return int(digitos.group())
        return 999999
    except:
        return 999999

def ordenar_logradouros_df(df_input):
    df_ordenado = df_input.copy()
    df_ordenado['_numero_ordem'] = df_ordenado['numero_logradouro'].apply(extrair_numero_logradouro)
    df_ordenado = df_ordenado.sort_values(
        by=['tipo_logradouro', 'nome_logradouro', '_numero_ordem', 'logradouro'],
        key=lambda x: x.str.lower() if x.dtype == "object" else x
    )
    df_ordenado = df_ordenado.drop(columns=['_numero_ordem'])
    return df_ordenado

# --- Fun√ß√£o Principal de L√≥gica ---

def execute_report_generator(processed_file_path, data_inicio, data_fim, log_callback):
    """
    Fun√ß√£o principal que executa toda a l√≥gica de gera√ß√£o de relat√≥rio.
    Recebe o caminho do arquivo processado, datas e uma fun√ß√£o de callback para o log.
    Retorna os caminhos dos arquivos gerados (planilha, relatorio_txt).
    """
    try:
        log_callback("=" * 80)
        log_callback("GERADOR DE RELAT√ìRIO CONSOLIDADO")
        log_callback("=" * 80)
        log_callback(f"‚úì Processamento iniciado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

        # Configura√ß√µes
        LIMIAR = 10
        script_dir = Path(__file__).parent
        projeto_root = script_dir
        
        if not (projeto_root / 'docs').exists():
             projeto_root = script_dir.parent
             if not (projeto_root / 'docs').exists():
                 log_callback(f"‚ùå Estrutura de pastas 'docs' n√£o encontrada a partir de {script_dir}")
                 raise FileNotFoundError("N√£o foi poss√≠vel localizar a pasta 'docs'")

        DOCS_DIR = projeto_root / 'docs'
        DOCS_DIR.mkdir(exist_ok=True)

        log_callback(f"\n‚úì Configura√ß√µes:")
        log_callback(f"  ‚Ä¢ Limiar de aglomera√ß√£o: > {LIMIAR} pessoas")
        log_callback(f"  ‚Ä¢ Diret√≥rio sa√≠da: {DOCS_DIR}")
        
        arquivo_selecionado = Path(processed_file_path)
        log_callback(f"‚úì Arquivo de entrada: {arquivo_selecionado.name}")
        log_callback(f"‚úì Per√≠odo definido: {data_inicio.strftime('%d/%m/%Y')} at√© {data_fim.strftime('%d/%m/%Y')}")

        # 6. Carregar e Preparar Dados
        log_callback(f"\nüìä Carregando dados...")
        df = pd.read_excel(arquivo_selecionado)
        log_callback(f"‚úì Planilha carregada: {len(df):,} registros")

        colunas_necessarias = ['Data', 'Per√≠odo', 'Qtd. pessoas', 'Logradouro', 'tipo_logradouro', 'nome_logradouro', 'numero_logradouro']
        colunas_faltantes = [col for col in colunas_necessarias if col not in df.columns]
        if colunas_faltantes:
            log_callback(f"\n‚ö†Ô∏è  ERRO: Colunas obrigat√≥rias n√£o encontradas: {colunas_faltantes}")
            raise KeyError(f"Colunas faltantes: {colunas_faltantes}")

        df = df.rename(columns={
            'Data': 'data', 'Per√≠odo': 'periodo', 'Qtd. pessoas': 'qtd_pessoas', 'Logradouro': 'logradouro'
        })
        df['data'] = pd.to_datetime(df['data'], errors='coerce')
        df = df.dropna(subset=['data'])
        df['qtd_pessoas'] = pd.to_numeric(df['qtd_pessoas'], errors='coerce')
        df = df.dropna(subset=['qtd_pessoas'])
        df['periodo_norm'] = df['periodo'].apply(normalizar_periodo)
        df = df.dropna(subset=['periodo_norm'])
        log_callback(f"‚úì Dados preparados")

        # 7. Calcular M√©dia Anterior
        if data_fim.weekday() == 0:  # 0 significa Segunda-feira
            dias_recuo = 3
            log_callback(f"üìÖ Relat√≥rio de Segunda-feira detectado: comparando com 3 dias atr√°s (Sexta-feira).")
        else:
            dias_recuo = 1
            
        data_inicio_anterior = data_inicio - timedelta(days=dias_recuo)
        data_fim_anterior = data_fim - timedelta(days=dias_recuo)
        
        df_anterior = df[(df['data'] >= data_inicio_anterior) & (df['data'] <= data_fim_anterior)].copy()

        media_anterior = 0.0
        if len(df_anterior) > 0:
            df_anterior['data_str'] = df_anterior['data'].dt.strftime('%d/%m/%Y')
            contagens_anterior = df_anterior.groupby(['logradouro', 'periodo_norm', 'data_str'])['qtd_pessoas'].sum().to_dict()
            dias_anteriores = gerar_lista_dias(data_inicio_anterior, data_fim_anterior)
            dias_validos_anterior = dias_anteriores[1:]
            dias_noite_anterior = dias_anteriores[:-1]
            df_logradouros_unicos_anterior = df_anterior[['logradouro', 'tipo_logradouro', 'nome_logradouro', 'numero_logradouro']].drop_duplicates()
            df_logradouros_ordenados_anterior = ordenar_logradouros_df(df_logradouros_unicos_anterior)
            logradouros_anterior = df_logradouros_ordenados_anterior['logradouro'].tolist()
            
            totais_por_coluna_anterior = []
            for periodo in ['madrugada', 'manh√£', 'tarde', 'noite']:
                dias_ref = dias_noite_anterior if periodo == 'noite' else dias_validos_anterior
                for dia in dias_ref:
                    dia_str = dia.strftime('%d/%m/%Y')
                    soma_coluna = sum(contagens_anterior.get((logradouro, periodo, dia_str), 0) for logradouro in logradouros_anterior)
                    if soma_coluna > 0:
                        totais_por_coluna_anterior.append(soma_coluna)
            
            if totais_por_coluna_anterior:
                media_anterior = round(sum(totais_por_coluna_anterior) / len(totais_por_coluna_anterior))
        
        if media_anterior > 0:
            log_callback(f"‚úì M√©dia anterior calculada: {media_anterior:.0f} pessoas/dia")
        else:
            log_callback(f"‚ö†Ô∏è  Sem dados para o intervalo anterior. M√©dia anterior = 0")

        df_periodo = df[(df['data'] >= data_inicio) & (df['data'] <= data_fim)].copy()
        log_callback(f"‚úì Dados do per√≠odo atual: {len(df_periodo):,} registros")

        # 8. Gerar Lista de Dias
        dias_lista = gerar_lista_dias(data_inicio, data_fim)
        dias_validos = dias_lista[1:]
        dias_noite = dias_lista[:-1]
        log_callback(f"‚úì Estrutura dos dias gerada.")

        # 9. Construir Matriz de Contagens
        log_callback(f"\nüîÑ Construindo matriz de contagens...")
        df_periodo['data_str'] = df_periodo['data'].dt.strftime('%d/%m/%Y')
        contagens = df_periodo.groupby(['logradouro', 'periodo_norm', 'data_str'])['qtd_pessoas'].sum().to_dict()
        df_logradouros_unicos = df_periodo[['logradouro', 'tipo_logradouro', 'nome_logradouro', 'numero_logradouro']].drop_duplicates()
        df_logradouros_ordenados = ordenar_logradouros_df(df_logradouros_unicos)
        logradouros = df_logradouros_ordenados['logradouro'].tolist()
        log_callback(f"‚úì {len(logradouros)} logradouros √∫nicos identificados e ordenados")

        # 10. Criar Cabe√ßalhos
        periodos = ['madrugada', 'manh√£', 'tarde', 'noite']
        periodos_fmt = {'madrugada': 'Madrugada', 'manh√£': 'Manh√£', 'tarde': 'Tarde', 'noite': 'Noite'}
        
        primeiro_dia = dias_validos[0] if dias_validos else data_inicio
        ultimo_dia = dias_validos[-1] if dias_validos else data_fim
        header1 = [f"Contagem di√°ria - Santa Cec√≠lia, Campos El√≠seos e Santa Ifig√™nia - {primeiro_dia.strftime('%d/%m/%Y')} a {ultimo_dia.strftime('%d/%m/%Y')}"]
        header2 = ['Ordem', 'Per√≠odo']
        header3 = ['', 'Logradouro' + ' ' * 20 + 'Data']

        for periodo in periodos:
            dias_ref = dias_noite if periodo == 'noite' else dias_validos
            for _ in dias_ref:
                header2.append(periodos_fmt[periodo])
            for dia in dias_ref:
                header3.append(dia.strftime('%d'))
        
        header2.extend(['M√©dia por per√≠odo', '', '', '', ''])
        header3.extend(['Madrugada', 'Manh√£', 'Tarde', 'Noite', '>10'])

        colunas_totais = len(header2)
        while len(header1) < colunas_totais: header1.append('')
        while len(header3) < colunas_totais: header3.append('')
        log_callback(f"‚úì Cabe√ßalhos criados: {colunas_totais} colunas")

        # 11. Construir Matriz de Dados
        log_callback(f"\nüîÑ Construindo matriz de dados...")
        matriz = []
        visiveis = []

        for logradouro in logradouros:
            linha = ['', logradouro]
            soma_linha = 0
            alguma_acima_limiar = False
            valores_por_periodo = {'madrugada': [], 'manh√£': [], 'tarde': [], 'noite': []}
            contador_acima_10 = 0
            
            for periodo in periodos:
                dias_ref = dias_noite if periodo == 'noite' else dias_validos
                for dia in dias_ref:
                    dia_str = dia.strftime('%d/%m/%Y')
                    chave = (logradouro, periodo, dia_str)
                    valor = contagens.get(chave, 0)
                    linha.append(valor if valor > 0 else '')
                    soma_linha += valor
                    if valor > LIMIAR:
                        alguma_acima_limiar = True
                        contador_acima_10 += 1
                    valores_por_periodo[periodo].append(valor)
            
            media_madrugada = round(sum(valores_por_periodo['madrugada']) / len(valores_por_periodo['madrugada'])) if valores_por_periodo['madrugada'] else ''
            media_manha = round(sum(valores_por_periodo['manh√£']) / len(valores_por_periodo['manh√£'])) if valores_por_periodo['manh√£'] else ''
            media_tarde = round(sum(valores_por_periodo['tarde']) / len(valores_por_periodo['tarde'])) if valores_por_periodo['tarde'] else ''
            media_noite = round(sum(valores_por_periodo['noite']) / len(valores_por_periodo['noite'])) if valores_por_periodo['noite'] else ''
            
            linha.extend([media_madrugada, media_manha, media_tarde, media_noite, contador_acima_10 if contador_acima_10 > 0 else ''])
            
            if soma_linha > 0:
                matriz.append(linha)
                visiveis.append(alguma_acima_limiar)

        ordem = 1
        for i in range(len(matriz)):
            if visiveis[i]:
                matriz[i][0] = ordem
                ordem += 1
            else:
                matriz[i][0] = ''
        log_callback(f"‚úì Matriz criada: {len(matriz)} logradouros")

        # 12. Calcular Linha de Totais
        total_row = [''] * colunas_totais
        total_row[1] = 'TOTAL'
        num_colunas_dados = colunas_totais - 5

        for col in range(2, num_colunas_dados):
            soma = sum(matriz[row][col] if isinstance(matriz[row][col], (int, float)) else 0 for row in range(len(matriz)))
            total_row[col] = soma if soma > 0 else ''

        total_valores_por_periodo = {'madrugada': [], 'manh√£': [], 'tarde': [], 'noite': []}
        col_idx = 2
        for periodo in periodos:
            dias_ref = dias_noite if periodo == 'noite' else dias_validos
            for _ in dias_ref:
                valor = total_row[col_idx]
                if isinstance(valor, (int, float)) and valor > 0:
                    total_valores_por_periodo[periodo].append(valor)
                col_idx += 1
        
        media_total_madr = round(sum(total_valores_por_periodo['madrugada']) / len(total_valores_por_periodo['madrugada'])) if total_valores_por_periodo['madrugada'] else ''
        media_total_manha = round(sum(total_valores_por_periodo['manh√£']) / len(total_valores_por_periodo['manh√£'])) if total_valores_por_periodo['manh√£'] else ''
        media_total_tarde = round(sum(total_valores_por_periodo['tarde']) / len(total_valores_por_periodo['tarde'])) if total_valores_por_periodo['tarde'] else ''
        media_total_noite = round(sum(total_valores_por_periodo['noite']) / len(total_valores_por_periodo['noite'])) if total_valores_por_periodo['noite'] else ''

        total_row[num_colunas_dados] = media_total_madr
        total_row[num_colunas_dados + 1] = media_total_manha
        total_row[num_colunas_dados + 2] = media_total_tarde
        total_row[num_colunas_dados + 3] = media_total_noite
        total_row[num_colunas_dados + 4] = ''

        valores_somados = [v for v in total_row[2:num_colunas_dados] if isinstance(v, (int, float)) and v > 0]
        media_atual = round(sum(valores_somados) / len(valores_somados)) if valores_somados else 0
        log_callback(f"\nüìä M√©dias calculadas:")
        log_callback(f"  ‚Ä¢ M√©dia atual: {media_atual:.0f} pessoas/dia")
        log_callback(f"  ‚Ä¢ M√©dia anterior: {media_anterior:.0f} pessoas/dia")

        # --- NOVO BLOCO: DETEC√á√ÉO E ORDENA√á√ÉO DE VARIA√á√ïES (M√çNIMO 10 PESSOAS) ---
        log_callback(f"üìù Detectando varia√ß√µes de volume >= 10 pessoas...")
        todas_variacoes = []
        DIFERENCA_MINIMA = 10 

        for logradouro in logradouros:
            for periodo in periodos:
                dias_ref = dias_noite if periodo == 'noite' else dias_validos
                for i in range(len(dias_ref) - 1):
                    v1 = contagens.get((logradouro, periodo, dias_ref[i].strftime('%d/%m/%Y')), 0)
                    v2 = contagens.get((logradouro, periodo, dias_ref[i+1].strftime('%d/%m/%Y')), 0)
                    
                    dif_bruta = v2 - v1
                    if abs(dif_bruta) >= DIFERENCA_MINIMA:
                        pct_info = (dif_bruta / v1 * 100) if v1 > 0 else 100.0
                        todas_variacoes.append({
                            'logradouro': logradouro, 'periodo': periodo,
                            'd1': dias_ref[i].strftime('%d/%m'), 'd2': dias_ref[i+1].strftime('%d/%m'),
                            'v1': v1, 'v2': v2, 'pct': pct_info, 'dif_bruta': dif_bruta
                        })
        
        # ORDENA√á√ÉO: Primeiro os maiores aumentos (desc), depois as maiores redu√ß√µes (asc)
        aumentos = sorted([v for v in todas_variacoes if v['dif_bruta'] > 0], key=lambda x: x['dif_bruta'], reverse=True)
        reducoes = sorted([v for v in todas_variacoes if v['dif_bruta'] < 0], key=lambda x: x['dif_bruta']) # Mais negativo primeiro
        variacoes_extremas = aumentos + reducoes
        # -----------------------------------------------------------------------

        # 13. Gerar Dados de An√°lise (C√°lculos)
        log_callback(f"\nüìù Gerando dados para o texto de an√°lise...")
        
        def somar_periodo_no_dia(periodo, dia_str):
            total, enderecos, soma_aglom = 0.0, 0, 0.0
            for logradouro in logradouros:
                valor = contagens.get((logradouro, periodo, dia_str), 0)
                total += valor
                if valor > LIMIAR:
                    enderecos += 1
                    soma_aglom += valor
            return {'total': total, 'enderecos': enderecos, 'soma_aglom': soma_aglom}

        ultimo_dia_val = dias_validos[-1] if dias_validos else data_fim
        ultimo_dia_noite = dias_noite[-1] if dias_noite else data_fim

        madr = somar_periodo_no_dia('madrugada', ultimo_dia_val.strftime('%d/%m/%Y'))
        manha = somar_periodo_no_dia('manh√£', ultimo_dia_val.strftime('%d/%m/%Y'))
        tarde = somar_periodo_no_dia('tarde', ultimo_dia_val.strftime('%d/%m/%Y'))
        noite = somar_periodo_no_dia('noite', ultimo_dia_noite.strftime('%d/%m/%Y'))

        ultimos_3_dias = dias_validos[-3:] if len(dias_validos) >= 3 else dias_validos
        soma_por_logradouro = {}
        for logradouro in logradouros:
            total = sum(contagens.get((logradouro, p, d.strftime('%d/%m/%Y')), 0) for p in periodos for d in ultimos_3_dias)
            if total > 0: soma_por_logradouro[logradouro] = total
        
        top_5_logradouros = sorted(soma_por_logradouro.items(), key=lambda x: x[1], reverse=True)[:5]
        variacao = round(((media_atual - media_anterior) / media_anterior) * 100, 1) if media_anterior > 0 else 0
        hoje = datetime.now()
        ref_texto = "sexta-feira" if hoje.weekday() == 0 else "ontem"

        # 14. Criar Rodap√©
        hoje_formatado = hoje.strftime('%d/%m/%Y')
        rodape = [
            ['Nota: As ruas sem aglomera√ß√£o (>10) no per√≠odo solicitado est√£o ocultas, mas constam na planilha.'],
            ['Fonte: SMS/Reden√ß√£o na Rua'],
            [f'Elaborado por: SGM/SEPE, em {hoje_formatado}']
        ]
        rodape_norm = [linha + [''] * (colunas_totais - len(linha)) for linha in rodape]
        log_callback(f"‚úì Rodap√© criado")

        # 15. Montar Sa√≠da Completa
        saida = [header1, header2, header3, *matriz, total_row, *rodape_norm]
        
        # 16. Exportar para Excel
        nome_arquivo_saida = f"relatorio_diario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        caminho_saida = DOCS_DIR / nome_arquivo_saida
        pd.DataFrame(saida).to_excel(caminho_saida, index=False, header=False, engine='openpyxl')

        # 17. Aplicar Formata√ß√£o
        log_callback(f"\nüé® Aplicando formata√ß√£o...")
        wb = load_workbook(caminho_saida)
        ws = wb.active
        fonte_bold, fonte_italic = Font(bold=True), Font(italic=True, size=10)
        fill_cinza = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        fill_azul = PatternFill(start_color='B7E1FA', end_color='B7E1FA', fill_type='solid')
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=colunas_totais)
        ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(1, 1).font, ws.cell(1, 1).fill = fonte_bold, fill_cinza
        
        for row in range(2, 4):
            for col in range(1, colunas_totais + 1):
                cell = ws.cell(row, col)
                cell.font, cell.fill, cell.border = fonte_bold, fill_cinza, border_thin
        
        ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
        ws.cell(2, 1).alignment = Alignment(horizontal='center', vertical='center')

        col_inicio = 3
        for periodo in periodos:
            qtd_dias = len(dias_noite) if periodo == 'noite' else len(dias_validos)
            if qtd_dias > 0:
                ws.merge_cells(start_row=2, start_column=col_inicio, end_row=2, end_column=col_inicio + qtd_dias - 1)
                ws.cell(2, col_inicio).alignment = Alignment(horizontal='center', vertical='center')
            col_inicio += qtd_dias

        col_medias = colunas_totais - 4
        ws.merge_cells(start_row=2, start_column=col_medias, end_row=2, end_column=col_medias + 3)
        ws.cell(2, col_medias).value = "M√©dia por per√≠odo"
        ws.cell(2, col_medias).alignment = Alignment(horizontal='center', vertical='center')
        
        col_maior10 = colunas_totais
        ws.merge_cells(start_row=2, start_column=col_maior10, end_row=3, end_column=col_maior10)
        ws.cell(2, col_maior10).value = ">10"
        ws.cell(2, col_maior10).alignment = Alignment(horizontal='center', vertical='center')

        primeira_linha_dados = 4
        col_madrugada = colunas_totais - 4
        for row_idx, visivel in enumerate(visiveis):
            row = primeira_linha_dados + row_idx
            if not visivel: ws.row_dimensions[row].hidden = True
            for col in range(1, colunas_totais + 1):
                cell = ws.cell(row, col)
                cell.alignment = Alignment(horizontal='center')
                cell.border = border_thin
                if col >= 3 and isinstance(cell.value, (int, float)) and cell.value > LIMIAR:
                     if col != col_maior10: cell.fill = fill_azul

        linha_total = primeira_linha_dados + len(matriz)
        for col in range(1, colunas_totais + 1):
            cell = ws.cell(linha_total, col)
            cell.font, cell.fill, cell.alignment, cell.border = fonte_bold, fill_cinza, Alignment(horizontal='center'), border_thin
            if col == col_maior10: cell.value = ''

        linha_rodape_inicio = linha_total + 1
        for row in range(linha_rodape_inicio, linha_rodape_inicio + 3):
            for col in range(1, colunas_totais + 1): ws.cell(row, col).font = fonte_italic

        linha_media = linha_rodape_inicio + 4
        ws.cell(linha_media, 2).value, ws.cell(linha_media, 2).font = 'M√©dia:', fonte_bold
        ws.cell(linha_media, 3).value, ws.cell(linha_media, 3).font = int(media_atual), fonte_bold
        ws.cell(linha_media, 3).alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width, ws.column_dimensions['B'].width = 8, 45
        for col in range(3, col_madrugada): ws.column_dimensions[get_column_letter(col)].width = 6
        for col in range(col_madrugada, col_maior10 + 1): ws.column_dimensions[get_column_letter(col)].width = 12

        wb.save(caminho_saida)
        log_callback(f"‚úì Formata√ß√£o aplicada")

        # 18. Exportar Texto de An√°lise
        log_callback(f"\nüìù Exportando texto de an√°lise...")
        nome_txt = f"{nome_arquivo_saida.replace('.xlsx', '')}_analise.txt"
        caminho_txt = DOCS_DIR / nome_txt

        report_data = {
            'data_inicio': data_inicio, 'data_fim': data_fim,
            'data_inicio_anterior': data_inicio_anterior, 'data_fim_anterior': data_fim_anterior,
            'hoje': hoje, 'media_atual': media_atual, 'media_anterior': media_anterior,
            'variacao': variacao, 'ultimo_dia_val': ultimo_dia_val, 'ultimo_dia_noite': ultimo_dia_noite,
            'madr': madr, 'manha': manha, 'tarde': tarde, 'noite': noite,
            'top_5_logradouros': top_5_logradouros, 
            'variacoes_extremas': variacoes_extremas, 
            'ref_texto': ref_texto
        }
        
        conteudo_txt = logic_text_generator.generate_analysis_text(report_data)
        with open(caminho_txt, 'w', encoding='utf-8') as f: f.write(conteudo_txt)

        # 19. Resumo Executivo
        log_callback(f"\n" + "=" * 80)
        log_callback("RESUMO EXECUTIVO")
        log_callback("=" * 80)
        log_callback(f"M√©dia Atual:    {int(media_atual)} pessoas/dia")
        log_callback(f"M√©dia Anterior: {int(media_anterior)} pessoas/dia")
        log_callback(f"Varia√ß√£o:       {variacao:+.1f}%")
        log_callback("-" * 40)
        log_callback(f"Destaques do dia {ultimo_dia_val.strftime('%d/%m/%Y')}:")
        log_callback(f"  ‚Ä¢ Madrugada: {int(madr['total'])} pessoas ({int(madr['enderecos'])} aglomera√ß√µes)")
        log_callback(f"  ‚Ä¢ Manh√£:     {int(manha['total'])} pessoas ({int(manha['enderecos'])} aglomera√ß√µes)")
        log_callback(f"  ‚Ä¢ Tarde:     {int(tarde['total'])} pessoas ({int(tarde['enderecos'])} aglomera√ß√µes)")
        log_callback(f"  ‚Ä¢ Noite:     {int(noite['total'])} pessoas ({int(noite['enderecos'])} aglomera√ß√µes)")
        log_callback("=" * 80)
        
        return str(caminho_saida), str(caminho_txt)

    except Exception as e:
        log_callback(f"\n‚ùå ERRO GERAL NO GERADOR DE RELAT√ìRIO ‚ùå")
        log_callback(traceback.format_exc())
        return None, None